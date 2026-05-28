import httpx
from bs4 import BeautifulSoup
import re
from datetime import datetime, timedelta, timezone
import time
import zipfile
import threading
import json
import os
import hashlib
import phonenumbers
from phonenumbers import geocoder
import requests
import signal
import sys
from langdetect import detect, LangDetectException, DetectorFactory
from colorama import init, Fore, Style
import qrcode
from io import BytesIO

# Real-time console output — nonaktifkan buffering stdout/stderr
sys.stdout.reconfigure(line_buffering=True)
sys.stderr.reconfigure(line_buffering=True)

def make_httpx_client(timeout=30):
    return httpx.Client(
        follow_redirects=True,
        timeout=timeout,
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36", "X-Requested-With": "XMLHttpRequest"}
    )

def make_requests_session():
    return requests.Session()

# ================= FILES =================
ACCOUNTS_FILE = "accounts.json"
COOKIES_FILE = "cookie.json"
CACHE_FILE = "file/sent_cache.json"
MAX_CACHE_SIZE = 2000
LANG_CODE_MAP = {  
    "id": "#Indonesia", "en": "#English", "fr": "#French", "es": "#Spanish",  
    "pt": "#Portuguese", "ar": "#Arabic", "ru": "#Russian", "tr": "#Turkish",  
    "hi": "#Hindi", "th": "#Thai", "vi": "#Vietnamese", "ms": "#Malay",  
    "tl": "#Filipino", "ja": "#Japanese", "ko": "#Korean", "zh-cn": "#Chinese",  
    "nl": "#Dutch", "sv": "#Swedish", "pl": "#Polish", "uk": "#Ukrainian",  
    "cs": "#Czech", "ro": "#Romanian", "el": "#Greek", "he": "#Hebrew", "fa": "#Persian"  
}  
    
# ================= CONFIG =================
OWNER_ID = 1611669051  # ID OWNER 
BASE = "https://ivaskicen2.serverkicen.biz.id"
LOGIN_URL = f"{BASE}/login"
RECV_URL  = f"{BASE}/portal/sms/received"          # Sumber recv_csrf (per-page CSRF iVAS)
GET_RANGE_URL = f"{BASE}/portal/sms/received/getsms"
GET_NUMBER_URL = f"{BASE}/portal/sms/received/getsms/number"
GET_SMS_URL = f"{BASE}/portal/sms/received/getsms/number/sms"
RETURN_NUMBER_URL = f"{BASE}/portal/numbers/return/number"
RETURN_ALL_URL = f"{BASE}/portal/numbers/return/allnumber/bluck"
EXPORT_URL = f"{BASE}/portal/numbers/export"

BOT_TOKEN = "8134173482:AAGHVHETip-dzRTrbPwvpiYPqX89p3N7rSc"
GROUPS_FILE = "groups.json"
ADDNUM_API_URL = "https://ws.websocket.web.id/admin/addnumber"
ADDNUM_API_KEY = "112231"
USERS_FILE = "users.json"
PREMIUM_FILE = "premium.json"
AMBIL_FILE = "file/ambil_nomor.json"
PREMIUM_COOKIE_FILE = "premium-cookie.json"
LINK_OWNER = "t.me/kicenxensai"
LINK_CHANNEL = "https://t.me/xorakuk"

# ================= LOG & FORCE JOIN =================
LOG_CHANNEL_ID = -1003908618331
BOT_USERNAME   = ""   # diisi otomatis saat startup via getMe

FORCE_JOIN_CHANNELS = [
    {"username": "numberchshiro", "url": "https://t.me/numberchshiro", "label": "📢 Channel Info Update"},
    {"username": "Ranzzz4",       "url": "https://t.me/Ranzzz4",       "label": "💬 Group Chat"},
]

# ================= PAKASIR PAYMENT =================
PAKASIR_PROJECT = os.getenv("PAKASIR_PROJECT", "kifzl")
PAKASIR_API_KEY  = os.getenv("PAKASIR_API_KEY", "pshebntafOuCoRpuQ5DBs2p9mML5cKU5")
PAKASIR_BASE     = "https://app.pakasir.com"

# Harga paket: {tier: {durasi_hari: harga_rupiah}}
PACKAGE_PRICES = {
    "starter": {1: 5_000,   7: 25_000,   30: 75_000},
    "pro":     {1: 10_000,  7: 50_000,   30: 150_000},
    "elite":   {1: 20_000,  7: 100_000,  30: 250_000},
    "ultra":   {1: 35_000,  7: 175_000,  30: 450_000},
}
DURATION_INFO = {
    1:  {"label": "Harian",   "emoji": "📅"},
    7:  {"label": "Mingguan", "emoji": "📆"},
    30: {"label": "Bulanan",  "emoji": "🗓️"},
}

SERVICE_SHORT = {
    "WHATSAPP": "#WS", "TELEGRAM": "#TG", "GOOGLE": "#G", "FACEBOOK": "#FB",
    "INSTAGRAM": "#IG", "SHOPEE": "#SP", "TOKOPEDIA": "#TP", "GRAB": "#GR",
    "GOJEK": "#GJ", "TIKTOK": "#TT"
}
sms_stats = {
    "total_sms": 0,
    "total_otp": 0,
    "total_number": set()
}
last_update_id = 0
MAX_EMAIL = 20 # Setting Max Email User/Owner
TOKEN_TIERS = {
    "free":    {"label": "FREE",    "emoji": "👤", "tokens_day": 20,    "max_email": 1},
    "starter": {"label": "STARTER", "emoji": "⭐", "tokens_day": 50,    "max_email": 3},
    "pro":     {"label": "PRO",     "emoji": "💎", "tokens_day": 150,   "max_email": 8},
    "elite":   {"label": "ELITE",   "emoji": "🔥", "tokens_day": 500,   "max_email": 15},
    "ultra":   {"label": "ULTRA",   "emoji": "👑", "tokens_day": 99999, "max_email": 20},
}
DetectorFactory.seed = 0
init(autoreset=True)
accounts_lock = threading.Lock()
LOGIN_COOLDOWN = 300  # 5 menit
SESSION_RETRY_INTERVAL = 600  # retry setiap 10 menit kalau session gagal

# ================= TELEGRAM SESSION (persistent + retry) =================
_TG_SESSION = requests.Session()
_TG_ADAPTER = requests.adapters.HTTPAdapter(
    pool_connections=4, pool_maxsize=10, max_retries=0
)
_TG_SESSION.mount("https://", _TG_ADAPTER)

def _tg_request(method, data=None, json_data=None, files=None, timeout=12):
    """
    Kirim request ke Telegram API via persistent session (connection pooling).
    Retry max 3x dengan exponential backoff + auto-handle 429 Flood Wait.
    """
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/{method}"
    for attempt in range(3):
        try:
            r = _TG_SESSION.post(url, data=data, json=json_data, files=files, timeout=timeout)
            if r.status_code == 429:
                retry_after = r.json().get("parameters", {}).get("retry_after", 5)
                print(Fore.YELLOW + f"  TG FLOOD 429 [{method}]: tunggu {retry_after}s")
                time.sleep(retry_after + 1)
                continue
            return r
        except Exception as e:
            if attempt == 2:
                print(Fore.RED + f"  TG ERROR [{method}]: {e}")
            else:
                time.sleep(1.5 ** (attempt + 1))
    return None

pending_setcookie = {}   # user_id -> {"email": str, "msg_id": int}
pending_addcookie = {}   # user_id -> {"email": str, "msg_id": int}
pending_addnum    = {}   # user_id -> {"email": str, "msg_id": int}
pending_payments  = {}   # user_id -> {order_id, tier, days, amount, chat_id, qr_msg_id}

# ================= SESSION TRACKER =================
_session_fail_time   = {}   # email -> timestamp pertama kali gagal
_session_notified    = {}   # email -> bool sudah notif atau belum
_session_retry_time  = {}   # email -> timestamp terakhir retry
_session_recovered   = {}   # email -> bool sudah notif recover

# ================= AUTO COOKIE REFRESHER =================
COOKIE_KEEPALIVE_INTERVAL = 600   # keepalive tiap 10 menit (sebelum session sempat expire)
COOKIE_NOTIF_COOLDOWN     = 3600  # notif ulang maks 1x per jam per akun
_last_cookie_refresh      = {}    # email -> timestamp terakhir keepalive
_last_cookie_notif        = {}    # email -> timestamp terakhir notif dikirim
_keepalive_warn_count     = {}    # email -> jumlah gagal keepalive berturut-turut

# ================= RANGES CACHE (kurangi beban IVAS server) =================
_ranges_cache    = {}   # email -> (timestamp, ranges_list)
RANGES_CACHE_TTL = 300  # 5 menit — ranges jarang berubah

# ================= RECV CSRF CACHE =================
# iVAS pakai per-page CSRF — /portal/sms/received punya token berbeda dari /portal
# Semua POST ke getsms, getsms/number, getsms/number/sms WAJIB pakai recv_csrf ini
_recv_csrf_cache = {}   # email -> {"csrf": str, "ts": float}
RECV_CSRF_TTL    = 900  # 15 menit — refresh sebelum expired

# ================= AUTO BACKUP =================
# Direktori & pola yang TIDAK perlu dibackup (sistem/cache/package)
BACKUP_SKIP_DIRS = {
    ".git", ".pythonlibs", ".local", ".cache",
    ".agents", ".upm", "nix", "__pycache__",
}
BACKUP_SKIP_EXTS = {".pyc", ".pyo", ".zip"}
BACKUP_SKIP_FILES = {".replit", "replit.nix"}

# ================= EXPIRY NOTIFIER =================
_notif_expiry_sent = {}   # str(uid) -> set {"24h", "3h", "1h"}

# ================= THREAD-SAFE CACHE & SHARED STATE =================
_sent_cache_lock  = threading.Lock()
_cache_dirty      = False
_last_cache_save  = 0.0
# Dibaca worker threads — diupdate oleh run_bot() manager thread
_bot_state = {"email_to_uid": {}, "total_accounts": 0}
# Flag untuk memaksa run_bot sync segera (set True setelah addcookie/setcookie berhasil)
_force_bot_sync   = False

# ================= ACCOUNT MANAGEMENT =================
def load_accounts():
    if not os.path.exists(ACCOUNTS_FILE):
        with open(ACCOUNTS_FILE, "w") as f:
            f.write('{"accounts":[]}')
    try:
        with open(ACCOUNTS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data.get("accounts", [])
    except:
        return []

def save_accounts():
    data_to_save = []
    for acc in accounts:
        data_to_save.append({
            "email": acc.get("email"),
            "password": acc.get("password"),
            "cookies": acc.get("cookies", {})
        })
    with open(ACCOUNTS_FILE, "w", encoding="utf-8") as f:
        json.dump({"accounts": data_to_save}, f, indent=2)

def load_cookies():
    if not os.path.exists(COOKIES_FILE):
        with open(COOKIES_FILE, "w") as f:
            f.write("{}")
        return {}
    try:
        with open(COOKIES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}

def load_premium():
    if not os.path.exists(PREMIUM_FILE):
        with open(PREMIUM_FILE, "w") as f:
            json.dump({}, f)
    try:
        with open(PREMIUM_FILE, "r") as f:
            return json.load(f)
    except:
        return {}

def save_premium(data):
    with open(PREMIUM_FILE, "w") as f:
        json.dump(data, f, indent=2)

premium_users = load_premium()

def get_user_tier(user_id):
    if user_id == OWNER_ID:
        return "ultra"
    user = premium_users.get(str(user_id))
    if not user:
        return "free"
    if time.time() > user.get("expired", 0):
        return "free"
    return user.get("tier", "free")

def get_tier_tokens_day(user_id):
    return TOKEN_TIERS.get(get_user_tier(user_id), TOKEN_TIERS["free"])["tokens_day"]

def get_tier_email_limit(user_id):
    if user_id == OWNER_ID:
        return MAX_EMAIL
    return TOKEN_TIERS.get(get_user_tier(user_id), TOKEN_TIERS["free"])["max_email"]

def is_premium(user_id):
    return get_user_tier(user_id) != "free"
    
def save_cookies(cookies_dict):
    with open(COOKIES_FILE, "w") as f:
        json.dump(cookies_dict, f, indent=2)

def extract_session_cookies(session):
    """Ekstrak semua cookies dari httpx session sebagai dict (fresh cookies)."""
    try:
        return dict(session.cookies)
    except:
        return {}

def save_fresh_cookies_auto(email, fresh_cookies):
    """Simpan fresh cookies ke file yang sesuai (owner/premium) berdasarkan email."""
    if not fresh_cookies:
        return
    all_cookies = load_cookies()
    if email in all_cookies:
        all_cookies[email] = fresh_cookies
        save_cookies(all_cookies)
        return
    prem = load_premium_cookies()
    if email in prem:
        prem[email] = fresh_cookies
        save_premium_cookies(prem)

def load_groups():
    if not os.path.exists(GROUPS_FILE):
        with open(GROUPS_FILE, "w") as f:
            json.dump({}, f)
    try:
        with open(GROUPS_FILE, "r") as f:
            data = json.load(f)
        # Backward compat: format lama {"groups": [...]} → convert ke dict kosong
        if isinstance(data, list) or "groups" in data:
            return {}
        return data
    except:
        return {}

def load_users():
    if not os.path.exists(USERS_FILE):
        with open(USERS_FILE, "w") as f:
            json.dump({}, f)
    try:
        with open(USERS_FILE, "r") as f:
            return json.load(f)
    except:
        return {}
        
def load_premium_cookies():
    if not os.path.exists(PREMIUM_COOKIE_FILE):
        with open(PREMIUM_COOKIE_FILE, "w") as f:
            json.dump({}, f)
    try:
        with open(PREMIUM_COOKIE_FILE, "r") as f:
            return json.load(f)
    except:
        return {}

def save_premium_cookies(data):
    with open(PREMIUM_COOKIE_FILE, "w") as f:
        json.dump(data, f, indent=2)

premium_cookies = load_premium_cookies()        

def save_users(data):
    with open(USERS_FILE, "w") as f:
        json.dump(data, f, indent=2)        

def save_groups():
    with open(GROUPS_FILE, "w") as f:
        json.dump(user_groups, f, indent=2)

user_groups = load_groups()

# ===== PER-USER GROUP HELPERS =====
def get_user_groups(user_id):
    return user_groups.get(str(user_id), [])

def add_user_group(user_id, gid):
    uid = str(user_id)
    if uid not in user_groups:
        user_groups[uid] = []
    if gid not in user_groups[uid]:
        user_groups[uid].append(gid)
    save_groups()

def remove_user_group(user_id, gid):
    uid = str(user_id)
    if uid in user_groups and gid in user_groups[uid]:
        user_groups[uid].remove(gid)
        save_groups()
        return True
    return False

# ===== USER KEY (IDENTITY) =====
def generate_user_key(user_id):
    raw = f"KICEN-{user_id}-IVAS-SECRET"
    h = hashlib.md5(raw.encode()).hexdigest()[:8].upper()
    return f"KX-{h[:4]}-{h[4:]}"

def get_or_create_user_key(user_id):
    if user_id == OWNER_ID:
        return generate_user_key(user_id)
    users = load_users()
    uid = str(user_id)
    changed = False
    if uid not in users:
        users[uid] = {"emails": [], "key": generate_user_key(user_id)}
        changed = True
    elif "key" not in users.get(uid, {}):
        users[uid]["key"] = generate_user_key(user_id)
        changed = True
    if changed:
        save_users(users)
    return users[uid]["key"]

# ================= TOKEN SYSTEM =================
TOKEN_MAX = 20  # token default free user per hari (reset jam 00:00 WIB)

def get_wib_date():
    """Return tanggal hari ini dalam WIB (UTC+7)."""
    from datetime import timezone, timedelta
    tz_wib = timezone(timedelta(hours=7))
    return datetime.now(tz_wib).strftime("%Y-%m-%d")

def get_user_tokens(user_id):
    """Ambil sisa token user hari ini. Auto-reset jika hari baru (WIB)."""
    if user_id == OWNER_ID:
        return 99999
    uid = str(user_id)
    users = load_users()
    today = get_wib_date()
    daily_limit = get_tier_tokens_day(user_id)
    if uid not in users:
        users[uid] = {"emails": [], "tokens": daily_limit, "last_token_reset": today}
        save_users(users)
        return daily_limit
    u = users[uid]
    if u.get("last_token_reset") != today:
        u["tokens"] = daily_limit
        u["last_token_reset"] = today
        save_users(users)
    return u.get("tokens", daily_limit)

def use_token(user_id):
    """Kurangi 1 token. Return True berhasil, False kalau habis. Owner unlimited."""
    if user_id == OWNER_ID:
        return True
    uid = str(user_id)
    users = load_users()
    today = get_wib_date()
    daily_limit = get_tier_tokens_day(user_id)
    if uid not in users:
        users[uid] = {"emails": [], "tokens": daily_limit, "last_token_reset": today}
    u = users[uid]
    if u.get("last_token_reset") != today:
        u["tokens"] = daily_limit
        u["last_token_reset"] = today
    if u.get("tokens", 0) <= 0:
        save_users(users)
        return False
    u["tokens"] = u.get("tokens", daily_limit) - 1
    save_users(users)
    return True

def token_status_str(user_id):
    """Return string singkat sisa token user untuk ditampilkan."""
    if user_id == OWNER_ID:
        return "♾️ Unlimited"
    t = get_user_tokens(user_id)
    daily = get_tier_tokens_day(user_id)
    return f"🎫 {t}/{daily}"

def no_token_msg(chat_id):
    send_msg(chat_id,
        "❌ <b>Token habis!</b>\n\n"
        "<blockquote>Token kamu sudah habis hari ini.\n"
        "Reset otomatis jam <b>00:00 WIB</b>.</blockquote>"
    )

# ===== PREMIUM ACCOUNT SESSION CACHE =====
_premium_acc_cache = {}  # email -> acc dict (persistent session untuk premium user)

# ===== USERNAME CACHE (in-memory, diisi saat user kirim pesan) =====
_username_cache = {}  # user_id (int) -> "@username" atau "uid:xxx"

_MAX_USERNAME_CACHE = 500

def store_username(user_id, from_obj):
    """Simpan username dari objek 'from' Telegram ke cache (max 500 entry)."""
    if len(_username_cache) >= _MAX_USERNAME_CACHE:
        # Hapus entry paling lama (FIFO)
        try:
            _username_cache.pop(next(iter(_username_cache)))
        except Exception:
            pass
    uname = from_obj.get("username")
    if uname:
        _username_cache[user_id] = f"@{uname}"
    else:
        first = from_obj.get("first_name", "")
        _username_cache[user_id] = first if first else f"uid:{user_id}"

def get_user_display(user_id):
    """Ambil label display user untuk console log."""
    return _username_cache.get(user_id, f"uid:{user_id}")

def load_sent_cache():
    os.makedirs("file", exist_ok=True)
    if not os.path.exists(CACHE_FILE):
        return set()
    try:
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return set(data) if isinstance(data, list) else set()
    except:
        return set()

def save_sent_cache():
    try:
        os.makedirs("file", exist_ok=True)
        cache_list = list(sent_cache)
        if len(cache_list) > MAX_CACHE_SIZE:
            cache_list = cache_list[-MAX_CACHE_SIZE:]
            sent_cache.clear()
            sent_cache.update(cache_list)
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache_list, f)
    except Exception as e:
        print(f"Error save cache: {e}")

def save_sent_cache_debounced():
    """Tandai cache perlu disimpan; flush max sekali per 5 detik."""
    global _cache_dirty, _last_cache_save
    _cache_dirty = True
    if time.time() - _last_cache_save >= 5:
        try:
            with _sent_cache_lock:
                save_sent_cache()
            _last_cache_save = time.time()
            _cache_dirty = False
        except Exception as e:
            print(f"WARN save cache: {e}")

# ================= LOAD DATA =================
accounts = load_accounts()
cookies_data = load_cookies()
sent_cache = load_sent_cache()

for acc in accounts:
    acc["session"] = make_httpx_client()
    acc["last_login"] = 0
    acc["csrf_token"] = "" 

    email = acc["email"]
    if email in cookies_data:
        acc["cookies"] = cookies_data[email]
        acc["session"].cookies.update(cookies_data[email])

# ================= ACCOUNT COMMANDS =================
def add_account(text):
    try:
        parts = text.split()
        if len(parts) < 3:
            tg_active("  Format: /addakun email password")
            return

        email, password = parts[1], parts[2]

        with accounts_lock:
            for acc in accounts:
                if acc["email"] == email:
                    tg_active(f"  Akun sudah ada: {email}")
                    return

            acc = {
                "email": email,
                "password": password,
                "cookies": {},
                "session": make_httpx_client(),
                "last_login": 0,
                "csrf_token": ""
            }

            accounts.append(acc)
            save_accounts()

        if login(acc):
            acc["last_login"] = time.time()
            tg_active(f"  Akun aktif & login: {email}")
        else:
            tg_active(f"   Akun masuk tapi login gagal: {email}")

    except Exception as e:
        tg_active(f"  Error add akun: {e}")
        
def save_number(number):
    try:
        with open(AMBIL_FILE, "r") as f:
            data = json.load(f)
    except:
        data = {"numbers": []}

    if number not in data["numbers"]:
        data["numbers"].append(number)

    with open(AMBIL_FILE, "w") as f:
        json.dump(data, f, indent=2)

def ambilnomor_to_txt():
    try:
        with open(AMBIL_FILE, "r") as f:
            data = json.load(f)
            numbers = data.get("numbers", [])
    except:
        numbers = []

    if not numbers:
        return None

    filename = "file/nomor.txt"
    with open(filename, "w") as f:
        for n in numbers:
            f.write(f"{n}\n") 

    return filename

def export_numbers_ivas(chat_id, acc, status_msg_id=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    email = acc["email"] if isinstance(acc, dict) else acc

    def _status(text):
        if status_msg_id:
            delete_and_send(chat_id, status_msg_id, text)
        else:
            send_msg(chat_id, text)

    # Ambil cookies dari acc atau fallback ke file
    if isinstance(acc, dict):
        cookies = acc.get("cookies") or {}
        if not cookies:
            all_cookies = load_cookies()
            prem_cookies = load_premium_cookies()
            cookies = all_cookies.get(email) or prem_cookies.get(email) or {}
    else:
        all_cookies = load_cookies()
        prem_cookies = load_premium_cookies()
        cookies = all_cookies.get(email) or prem_cookies.get(email) or {}

    if not cookies:
        _status(
            f"📁 <b>AMBIL FILE</b>\n\n"
            f"<blockquote>"
            f"📧 Email: <code>{email}</code>\n"
            f"❌ Cookie tidak ditemukan. Set cookie dulu."
            f"</blockquote>")
        return

    session = make_requests_session()
    session.cookies.update(cookies)

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": f"{BASE}/portal/numbers",
    }

    dt_params_base = {
        "columns[0][data]": "number_id",  "columns[0][name]": "id",
        "columns[1][data]": "Number",      "columns[1][name]": "Number",
        "columns[2][data]": "range",       "columns[2][name]": "range",
        "columns[3][data]": "A2P",         "columns[3][name]": "A2P",
        "columns[4][data]": "LimitA2P",    "columns[4][name]": "LimitA2P",
        "columns[5][data]": "limit_cli_a2p",     "columns[5][name]": "limit_cli_a2p",
        "columns[6][data]": "limit_cli_did_a2p", "columns[6][name]": "limit_cli_did_a2p",
        "columns[7][data]": "action",      "columns[7][name]": "action",
        "order[0][column]": 1, "order[0][dir]": "desc",
        "search[value]": "", "search[regex]": "false",
    }

    try:
        # Cek login + ambil total records
        r_check = session.get(f"{BASE}/portal/numbers", params={**dt_params_base, "draw": 1, "start": 0, "length": 1},
                              headers=headers, timeout=30)

        if "/login" in r_check.url:
            _status(
                f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Gagal login/verifikasi session. Perbarui cookie."
                f"</blockquote>")
            return

        if r_check.status_code != 200:
            _status(
                f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Server mengembalikan error: HTTP {r_check.status_code}\n"
                f"📄 Respon: <code>{r_check.text[:200]}</code>"
                f"</blockquote>")
            return

        try:
            meta = r_check.json()
        except Exception:
            _status(
                f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Gagal parse respon server. Cookie mungkin expired.\n"
                f"📄 Respon: <code>{r_check.text[:200]}</code>"
                f"</blockquote>")
            return

        total = meta.get("recordsTotal", 0)
        if total == 0:
            _status(
                f"📁 <b>AMBIL FILE</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"⚠️ Tidak ada nomor aktif untuk di-export."
                f"</blockquote>")
            return

        # Ambil semua data
        r_all = session.get(f"{BASE}/portal/numbers",
                            params={**dt_params_base, "draw": 2, "start": 0, "length": total},
                            headers=headers, timeout=120)

        if r_all.status_code != 200:
            _status(
                f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Server error saat ambil data: HTTP {r_all.status_code}\n"
                f"📄 Respon: <code>{r_all.text[:200]}</code>"
                f"</blockquote>")
            return

        try:
            data = r_all.json().get("data", [])
        except Exception:
            _status(
                f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Gagal parse data dari server.\n"
                f"📄 Respon: <code>{r_all.text[:200]}</code>"
                f"</blockquote>")
            return

        if not data:
            _status(
                f"📁 <b>AMBIL FILE</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"⚠️ Data kosong, coba lagi."
                f"</blockquote>")
            return

        # Build Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Numbers"

        header_row = ["No", "Number", "Range", "Rate (A2P)", "Limit by Range", "SID/Range Limit", "SID→DID Limit"]
        header_fill = PatternFill("solid", fgColor="2D6A9F")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(header_row, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for i, row in enumerate(data, 1):
            num_raw = row.get("Number", "")
            try:
                num_str = str(int(num_raw))
            except Exception:
                num_str = str(num_raw)

            ws.append([
                i,
                num_str,
                row.get("range", ""),
                row.get("A2P", ""),
                row.get("LimitA2P", ""),
                row.get("limit_cli_a2p", ""),
                row.get("limit_cli_did_a2p", ""),
            ])

        # Auto-width kolom
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{now}_ivas_numbers_{email.split('@')[0]}.xlsx"
        os.makedirs("file", exist_ok=True)
        filepath = f"file/{filename}"
        wb.save(filepath)

        if status_msg_id:
            delete_msg(chat_id, status_msg_id)

        with open(filepath, "rb") as f:
            requests.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                data={
                    "chat_id": chat_id,
                    "caption": (
                        f"📁 <b>FILE IVAS BERHASIL DIAMBIL</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email  : <code>{email}</code>\n"
                        f"🔢 Total  : <b>{len(data)}</b> nomor\n"
                        f"📄 File   : <code>{filename}</code>\n"
                        f"🕐 Waktu  : <code>{now}</code>\n"
                        f"✅ Status : Berhasil"
                        f"</blockquote>"
                    ),
                    "parse_mode": "HTML"
                },
                files={"document": (filename, f)}
            )
        os.remove(filepath)

    except Exception as e:
        _status(
            f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
            f"<blockquote>"
            f"📧 Email: <code>{email}</code>\n"
            f"❌ Error: <code>{str(e)[:200]}</code>"
            f"</blockquote>")
        
def del_account(text):
    try:
        _, email = text.split()
        global accounts
        accounts = [a for a in accounts if a["email"] != email]
        save_accounts()
        tg_active(f"  Akun dihapus: {email}")
    except:
        tg_active("  Format salah /delakun email")

def detect_language(text):
    try:
        if not text or len(text) < 10: return "#Unknown"
        text = re.sub(r"\d+", "", text).strip()
        if len(text) < 5: return "#Unknown"
        lang_code = detect(text)
        return LANG_CODE_MAP.get(lang_code, f"#{lang_code.upper()}")
    except LangDetectException:
        return "#Unknown"
        
def list_accounts(chat_id, user_id):
    try:
        if not accounts:
            send_msg(chat_id, "Belum ada akun")
            return
        msg = "  <b>LIST AKUN</b>\n\n"
        now = time.time()
        for i, acc in enumerate(accounts, 1):
            email = acc.get("email", "Unknown")
            safe_email = email if user_id == OWNER_ID else mask_email(email)
            last_login = acc.get("last_login", 0)
            status = "ACTIVE  " if now - last_login < LOGIN_COOLDOWN else "OFFLINE  "
            msg += f"{i}. {safe_email} | {status}\n"
        send_msg(chat_id, msg)
    except Exception as e:
        send_msg(chat_id, f"  Error list akun: {e}")
        
def add_token_tier(text, chat_id):
    """Owner: /addtoken user_id tier hari"""
    try:
        parts = text.split()
        if len(parts) < 4:
            return send_msg(chat_id,
                "❌ Format:\n<code>/addtoken user_id tier hari</code>\n\n"
                "Tier: <b>starter / pro / elite / ultra</b>\n"
                "Contoh: <code>/addtoken 123456789 pro 30</code>")
        uid, tier, hari = parts[1], parts[2].lower(), int(parts[3])
        if tier not in TOKEN_TIERS or tier == "free":
            return send_msg(chat_id, "❌ Tier tidak valid!\nPilih: starter / pro / elite / ultra")
        expired = time.time() + (hari * 86400)
        premium_users[str(uid)] = {"tier": tier, "expired": expired}
        save_premium(premium_users)
        t = TOKEN_TIERS[tier]
        tok_str = "♾️ Unlimited" if t["tokens_day"] >= 99999 else f"{t['tokens_day']}/hari"
        send_msg(chat_id,
            f"✅ <b>PAKET TOKEN AKTIF</b>\n\n"
            f"<blockquote>"
            f"👤 User ID  : <code>{uid}</code>\n"
            f"🏷️ Paket   : {t['emoji']} <b>{t['label']}</b>\n"
            f"🎫 Token   : {tok_str}\n"
            f"📧 Max Email: {t['max_email']} akun\n"
            f"📅 Durasi  : {hari} hari"
            f"</blockquote>")
        try:
            send_msg(int(uid),
                f"🎉 <b>PAKET TOKEN AKTIF!</b>\n\n"
                f"<blockquote>"
                f"🏷️ Paket    : {t['emoji']} <b>{t['label']}</b>\n"
                f"🎫 Token    : {tok_str}\n"
                f"📧 Max Email: {t['max_email']} akun\n"
                f"📅 Durasi   : {hari} hari\n\n"
                f"Token reset otomatis jam 00:00 WIB.\n"
                f"Ketik /cekprem untuk cek status."
                f"</blockquote>")
        except: pass
    except Exception as e:
        send_msg(chat_id, f"❌ Error: {e}")
        
def add_cookie_premium(text, chat_id, user_id):
    cmd_addcookie(chat_id, user_id)  
        
def del_cookie_premium(text, chat_id, user_id):
    try:
        parts = text.split()
        if len(parts) < 2:
            return send_msg(chat_id, "❌ Format:\n/delcookie email@gmail.com")
        email = parts[1].strip().lower()

        # Cek kepemilikan: email harus milik user ini (atau owner)
        if not is_owner(user_id):
            users_d = load_users()
            owned = users_d.get(str(user_id), {}).get("emails", [])
            if email not in owned:
                return send_msg(chat_id, "❌ Akun tidak ditemukan di akun kamu")

        # Hapus dari premium-cookie.json
        premium_cookies = load_premium_cookies()
        if email in premium_cookies:
            del premium_cookies[email]
            save_premium_cookies(premium_cookies)

        # Hapus dari user_accounts (password-based) di users.json
        users_d = load_users()
        uid = str(user_id)
        if uid in users_d:
            before = users_d[uid].get("user_accounts", [])
            users_d[uid]["user_accounts"] = [a for a in before if a.get("email") != email]
            # Hapus juga dari emails list
            if email in users_d[uid].get("emails", []):
                users_d[uid]["emails"].remove(email)
            save_users(users_d)

        send_msg(chat_id, f"✅ Akun dihapus:\n<code>{email}</code>")
    except Exception as e:
        send_msg(chat_id, f"❌ Error: {e}")                    
        
def del_token_tier(text, chat_id):
    """Owner: /deltoken user_id"""
    try:
        _, uid = text.split()
        if uid not in premium_users:
            return send_msg(chat_id, "❌ User tidak memiliki paket aktif")
        del premium_users[uid]
        save_premium(premium_users)
        send_msg(chat_id, f"✅ Paket user <code>{uid}</code> dihapus → kembali ke FREE")
        try:
            send_msg(int(uid), "⚠️ Paket token kamu telah dinonaktifkan oleh owner.")
        except: pass
    except:
        send_msg(chat_id, "❌ Format:\n/deltoken user_id")

def is_owner(user_id): return user_id == OWNER_ID

def list_token_tier(chat_id):
    """Owner: /listtoken"""
    if not premium_users:
        return send_msg(chat_id, "Belum ada user dengan paket aktif.")
    now = time.time()
    msg = "🏆 <b>LIST PAKET TOKEN AKTIF</b>\n\n"
    for i, (uid, data) in enumerate(premium_users.items(), 1):
        tier = data.get("tier", "starter")
        t = TOKEN_TIERS.get(tier, TOKEN_TIERS["starter"])
        sisa = max(0, int((data.get("expired", 0) - now) // 86400))
        msg += f"{i}. <code>{uid}</code> | {t['emoji']} {t['label']} | {sisa} hari\n"
    send_msg(chat_id, msg)

def cmd_beli(chat_id, user_id):
    tier = get_user_tier(user_id)
    t_cur = TOKEN_TIERS.get(tier, TOKEN_TIERS["free"])
    aktif_tag = f"  ✅ aktif" if tier != "free" else ""
    msg = (
        "🛒 <b>PAKET TOKEN IVAS</b>\n\n"
        "<blockquote>"
        "Pilih paket untuk melihat detail benefit.\n"
        f"Paket kamu saat ini: {t_cur['emoji']} <b>{t_cur['label']}</b>{aktif_tag}"
        "</blockquote>\n\n"
        "👇 Pilih paket:"
    )
    rows = [
        [
            {"text": "⭐ STARTER", "callback_data": "pkg_info:starter"},
            {"text": "💎 PRO",     "callback_data": "pkg_info:pro"},
        ],
        [
            {"text": "🔥 ELITE",   "callback_data": "pkg_info:elite"},
            {"text": "👑 ULTRA",   "callback_data": "pkg_info:ultra"},
        ]
    ]
    send_inline_keyboard_grid(chat_id, msg, rows)

def handle_pkg_info_cb(chat_id, user_id, tier_key, cb_id, msg_id):
    answer_callback_query(cb_id)
    if tier_key not in TOKEN_TIERS or tier_key == "free":
        return
    t = TOKEN_TIERS[tier_key]
    tok_str = "♾️ Unlimited" if t["tokens_day"] >= 99999 else f"{t['tokens_day']} token/hari"
    tier_now = get_user_tier(user_id)
    aktif_label = " ✅ <i>(paket kamu sekarang)</i>" if tier_now == tier_key else ""
    prices = PACKAGE_PRICES.get(tier_key, {})

    price_lines = ""
    for days, info in DURATION_INFO.items():
        harga = prices.get(days, 0)
        price_lines += f"  {info['emoji']} {info['label']:9s}: <b>Rp {harga:,}</b>\n".replace(",", ".")

    detail = (
        f"{t['emoji']} <b>PAKET {t['label']}</b>{aktif_label}\n\n"
        f"<blockquote>"
        f"🎫 Token/hari  : <b>{tok_str}</b>\n"
        f"📧 Max Email   : <b>{t['max_email']} akun</b> IVAS\n"
        f"🔄 Reset Token : 00:00 WIB\n\n"
        f"✅ <b>Semua fitur aktif:</b>\n"
        f"  • /addcookie — kelola cookie IVAS\n"
        f"  • /addemail — tambah akun IVAS\n"
        f"  • /addnum — tambah nomor test\n"
        f"  • /delnumall — kembalikan semua nomor\n"
        f"  • /myrange — cek range aktif\n"
        f"  • /ambilfile — export nomor ke Excel\n"
        f"  • /cekivas — statistik IVAS\n"
        f"  • SMS notif otomatis ke grup/PM\n\n"
        f"💰 <b>Harga:</b>\n"
        f"{price_lines}"
        f"</blockquote>\n\n"
        f"👇 Pilih durasi untuk lanjut bayar via QRIS:"
    )
    rows = []
    for days, info in DURATION_INFO.items():
        harga = prices.get(days, 0)
        rows.append([{
            "text": f"{info['emoji']} {info['label']} — Rp {harga:,}".replace(",", "."),
            "callback_data": f"pkg_buy:{tier_key}:{days}"
        }])
    rows.append([{"text": "🔙 Lihat Paket Lain", "callback_data": "pkg_back"}])
    delete_msg(chat_id, msg_id)
    send_inline_keyboard_grid(chat_id, detail, rows)


# ================= PAKASIR PAYMENT FUNCTIONS =================

def pakasir_create_qris(order_id, amount):
    try:
        r = requests.post(
            f"{PAKASIR_BASE}/api/transactioncreate/qris",
            json={"project": PAKASIR_PROJECT, "order_id": order_id,
                  "amount": amount, "api_key": PAKASIR_API_KEY},
            timeout=15
        )
        data = r.json()
        return data.get("payment")
    except Exception as e:
        print(f"[PAKASIR] create error: {e}")
        return None

def pakasir_check_status(order_id, amount):
    try:
        r = requests.get(
            f"{PAKASIR_BASE}/api/transactiondetail",
            params={"project": PAKASIR_PROJECT, "order_id": order_id,
                    "amount": amount, "api_key": PAKASIR_API_KEY},
            timeout=10
        )
        data = r.json()
        return data.get("transaction", {}).get("status", "unknown")
    except Exception as e:
        print(f"[PAKASIR] status error: {e}")
        return "unknown"

def pakasir_cancel(order_id, amount):
    try:
        requests.post(
            f"{PAKASIR_BASE}/api/transactioncancel",
            json={"project": PAKASIR_PROJECT, "order_id": order_id,
                  "amount": amount, "api_key": PAKASIR_API_KEY},
            timeout=10
        )
    except Exception as e:
        print(f"[PAKASIR] cancel error: {e}")

def generate_qr_image(qr_string):
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M,
                       box_size=10, border=4)
    qr.add_data(qr_string)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf

def send_photo_msg(chat_id, photo_bytes, caption, reply_markup=None):
    data = {"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"}
    if reply_markup:
        data["reply_markup"] = json.dumps(reply_markup)
    r = requests.post(
        f"https://api.telegram.org/bot{BOT_TOKEN}/sendPhoto",
        data=data,
        files={"photo": ("qris.png", photo_bytes, "image/png")},
        timeout=20
    )
    res = r.json()
    if res.get("ok"):
        return res["result"]["message_id"]
    return None

def payment_checker(user_id, chat_id, order_id, tier, days, amount, qr_msg_id):
    deadline = time.time() + 15 * 60  # 15 menit
    t = TOKEN_TIERS.get(tier, {})
    label = t.get("label", tier.upper())
    while time.time() < deadline:
        time.sleep(5)
        if pending_payments.get(user_id, {}).get("order_id") != order_id:
            return
        status = pakasir_check_status(order_id, amount)
        if status == "completed":
            pending_payments.pop(user_id, None)
            delete_msg(chat_id, qr_msg_id)
            add_token_tier(user_id, tier, days)
            emoji = t.get("emoji", "✅")
            exp_str = get_tier_expiry_str(user_id)
            send_msg(chat_id,
                f"✅ <b>PEMBAYARAN BERHASIL!</b>\n\n"
                f"Paket {emoji} <b>{label}</b> selama <b>{days} hari</b> telah aktif.\n"
                f"📅 Aktif hingga: <b>{exp_str}</b>\n\n"
                f"Selamat menggunakan bot IVAS! 🚀"
            )
            # Kirim laporan pembelian ke channel log
            def _purchase_log():
                uname = BOT_USERNAME
                now_str = datetime.now(timezone(timedelta(hours=7))).strftime("%d/%m/%Y %H:%M")
                udisp_buy = get_user_display(user_id)
                msg_log = (
                    f"💰 <b>PEMBELIAN BERHASIL</b>\n\n"
                    f"<blockquote>"
                    f"👤 User    : {udisp_buy}\n"
                    f"🆔 ID      : <code>{user_id}</code>\n"
                    f"📦 Paket   : {emoji} <b>{label}</b>\n"
                    f"📅 Durasi  : <b>{days} hari</b>\n"
                    f"💵 Nominal : <b>Rp {amount:,}</b>\n".replace(",", ".") +
                    f"🕐 Waktu   : {now_str} WIB\n"
                    f"🔖 Order   : <code>{order_id}</code>"
                    f"</blockquote>"
                )
                markup = {"inline_keyboard": [[{"text": "🤖 Buka Bot", "url": f"https://t.me/{uname}"}]]} if uname else None
                try:
                    payload = {"chat_id": LOG_CHANNEL_ID, "text": msg_log, "parse_mode": "HTML"}
                    if markup:
                        payload["reply_markup"] = markup
                    requests.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage", json=payload, timeout=10)
                except Exception as le:
                    print(f"PURCHASE LOG ERROR: {le}")
            threading.Thread(target=_purchase_log, daemon=True).start()
            return
        elif status in ("expired", "cancelled"):
            pending_payments.pop(user_id, None)
            delete_msg(chat_id, qr_msg_id)
            send_msg(chat_id,
                f"⏰ <b>Pembayaran kadaluarsa.</b>\n\n"
                f"QR QRIS untuk paket <b>{label}</b> sudah tidak valid.\n"
                f"Ketik /beli untuk membuat tagihan baru."
            )
            return
    pending_payments.pop(user_id, None)
    delete_msg(chat_id, qr_msg_id)
    pakasir_cancel(order_id, amount)
    send_msg(chat_id,
        f"⏰ <b>Waktu bayar habis (15 menit).</b>\n\n"
        f"Tagihan paket <b>{label}</b> dibatalkan otomatis.\n"
        f"Ketik /beli untuk membuat tagihan baru."
    )

def get_tier_expiry_str(user_id):
    premium = load_premium()
    entry = premium.get(str(user_id), {})
    exp = entry.get("expired", "")
    if not exp:
        return "-"
    try:
        dt = datetime.fromisoformat(exp)
        return dt.strftime("%d %b %Y %H:%M WIB")
    except Exception:
        return exp

def handle_pkg_buy_cb(chat_id, user_id, data, cb_id, msg_id):
    answer_callback_query(cb_id)
    parts = data.split(":")
    if len(parts) != 2:
        return
    tier_key, days_str = parts
    try:
        days = int(days_str)
    except ValueError:
        return
    if tier_key not in PACKAGE_PRICES or days not in PACKAGE_PRICES[tier_key]:
        send_msg(chat_id, "❌ Paket tidak valid.")
        return
    if not PAKASIR_PROJECT or not PAKASIR_API_KEY:
        send_msg(chat_id, "❌ Pembayaran belum dikonfigurasi. Hubungi owner.")
        return
    if user_id in pending_payments:
        old = pending_payments[user_id]
        send_msg(chat_id,
            f"⚠️ Kamu masih punya tagihan aktif (Order <code>{old['order_id']}</code>).\n"
            f"Selesaikan dulu atau batalkan dengan tombol di bawah QR sebelumnya."
        )
        return

    amount = PACKAGE_PRICES[tier_key][days]
    t = TOKEN_TIERS.get(tier_key, {})
    label = t.get("label", tier_key.upper())
    dur_info = DURATION_INFO[days]
    ts = int(time.time())
    order_id = f"IVAS{user_id}{ts}"

    loading_msg = send_msg_return_id(chat_id,
        f"⏳ Membuat tagihan QRIS untuk paket <b>{t.get('emoji','')} {label}</b> "
        f"{dur_info['label']}...\nMohon tunggu sebentar."
    )

    payment = pakasir_create_qris(order_id, amount)
    if not payment:
        if loading_msg:
            delete_msg(chat_id, loading_msg)
        send_msg(chat_id, "❌ Gagal membuat tagihan. Coba lagi atau hubungi owner.")
        return

    qr_string   = payment.get("payment_number", "")
    total        = payment.get("total_payment", amount)
    expired_at   = payment.get("expired_at", "")
    try:
        exp_dt = datetime.fromisoformat(expired_at.replace("Z", "+00:00"))
        exp_str = exp_dt.strftime("%H:%M WIB, %d %b %Y")
    except Exception:
        exp_str = expired_at

    if loading_msg:
        delete_msg(chat_id, loading_msg)

    try:
        qr_buf = generate_qr_image(qr_string)
    except Exception as e:
        print(f"[QR] generate error: {e}")
        send_msg(chat_id, "❌ Gagal membuat QR code. Coba lagi.")
        return

    caption = (
        f"📲 <b>TAGIHAN QRIS — {t.get('emoji','')} {label} {dur_info['label']}</b>\n\n"
        f"<blockquote>"
        f"💰 Total Bayar : <b>Rp {total:,}</b>\n".replace(",", ".") +
        f"⏰ Berlaku s/d : <b>{exp_str}</b>\n\n"
        f"📋 Order ID    : <code>{order_id}</code>"
        f"</blockquote>\n\n"
        f"Scan QR di atas menggunakan aplikasi e-wallet / m-banking.\n"
        f"Paket aktif <b>otomatis</b> setelah pembayaran berhasil. ✅"
    )
    markup = {"inline_keyboard": [[
        {"text": "❌ Batalkan Pembayaran", "callback_data": f"cancel_payment:{order_id}:{amount}"}
    ]]}
    qr_msg_id = send_photo_msg(chat_id, qr_buf, caption, markup)
    if not qr_msg_id:
        send_msg(chat_id, "❌ Gagal mengirim QR. Coba lagi.")
        return

    pending_payments[user_id] = {
        "order_id": order_id, "tier": tier_key, "days": days,
        "amount": amount, "chat_id": chat_id, "qr_msg_id": qr_msg_id
    }
    delete_msg(chat_id, msg_id)
    threading.Thread(
        target=payment_checker,
        args=(user_id, chat_id, order_id, tier_key, days, amount, qr_msg_id),
        daemon=True
    ).start()

def send_msg_return_id(chat_id, text):
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
            timeout=10
        )
        res = r.json()
        if res.get("ok"):
            return res["result"]["message_id"]
    except Exception:
        pass
    return None

def send_msg(chat_id, text):
    _tg_request("sendMessage", data={"chat_id": chat_id, "text": text, "parse_mode": "HTML"})

# ================= FORCE JOIN =================
def check_force_join(user_id):
    """Cek apakah user belum join channel/grup wajib. Return list yang belum."""
    not_joined = []
    for ch in FORCE_JOIN_CHANNELS:
        try:
            r = _tg_request("getChatMember", data={
                "chat_id": f"@{ch['username']}",
                "user_id": user_id
            })
            if r:
                res = r.json()
                if res.get("ok"):
                    status = res["result"]["status"]
                    if status in ("member", "administrator", "creator"):
                        continue
        except Exception:
            pass
        not_joined.append(ch)
    return not_joined

def send_force_join_msg(chat_id, not_joined):
    """Kirim pesan wajib join dengan tombol URL (bukan link teks)."""
    rows = [[{"text": ch["label"], "url": ch["url"]}] for ch in not_joined]
    rows.append([{"text": "✅ Saya Sudah Join", "callback_data": "check_join"}])
    keyboard = {"inline_keyboard": rows}
    text = (
        "🚫 <b>WAJIB JOIN DULU!</b>\n\n"
        "<blockquote>Kamu belum join semua channel/grup yang diwajibkan.\n"
        "Join dulu, lalu klik <b>✅ Saya Sudah Join</b>.</blockquote>"
    )
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML", "reply_markup": keyboard},
            timeout=10
        )
    except Exception as e:
        print(f"FORCE JOIN MSG ERROR: {e}")

# ================= ACTIVITY LOG =================
def send_activity_log(user_id, username_display, feature_name, status="✅ Berhasil"):
    """Kirim laporan aktivitas ke channel log secara async (tidak blocking)."""
    def _send():
        now_str = datetime.now(timezone(timedelta(hours=7))).strftime("%d/%m/%Y %H:%M")
        msg = (
            f"📋 <b>LAPORAN AKTIVITAS</b>\n\n"
            f"<blockquote>"
            f"👤 User   : {username_display}\n"
            f"🆔 ID     : <code>{user_id}</code>\n"
            f"🔧 Fitur  : <b>{feature_name}</b>\n"
            f"📊 Status : {status}\n"
            f"🕐 Waktu  : {now_str} WIB"
            f"</blockquote>"
        )
        uname = BOT_USERNAME
        markup = {"inline_keyboard": [[{"text": "🤖 Buka Bot", "url": f"https://t.me/{uname}"}]]} if uname else None
        try:
            payload = {"chat_id": LOG_CHANNEL_ID, "text": msg, "parse_mode": "HTML"}
            if markup:
                payload["reply_markup"] = markup
            requests.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                json=payload, timeout=10
            )
        except Exception as e:
            print(f"ACTIVITY LOG ERROR: {e}")
    threading.Thread(target=_send, daemon=True).start()
    
def cek_premium(chat_id, user_id):
    my_groups = get_user_groups(user_id)
    grup_status = f"{len(my_groups)} grup" if my_groups else "Belum addgrup (PM aktif)"
    tok = token_status_str(user_id)
    users_d = load_users()
    email_count = len(users_d.get(str(user_id), {}).get("emails", []))
    tier = get_user_tier(user_id)
    t = TOKEN_TIERS.get(tier, TOKEN_TIERS["free"])

    if user_id == OWNER_ID:
        return send_msg(chat_id,
            f"📊 <b>STATUS AKUN</b>\n\n"
            f"<blockquote>"
            f"👑 Mode     : OWNER\n"
            f"🎫 Token    : {tok}\n"
            f"📧 Email    : {email_count} akun\n"
            f"💬 Grup     : {grup_status}"
            f"</blockquote>")

    prem = premium_users.get(str(user_id))
    if prem and time.time() > prem.get("expired", 0):
        del premium_users[str(user_id)]
        save_premium(premium_users)
        prem = None
        tier = "free"
        t = TOKEN_TIERS["free"]

    user_key = get_or_create_user_key(user_id)
    email_limit = get_tier_email_limit(user_id)

    if tier == "free":
        send_msg(chat_id,
            f"📊 <b>STATUS AKUN</b>\n\n"
            f"<blockquote>"
            f"{t['emoji']} Paket    : <b>{t['label']}</b>\n"
            f"🔑 Key      : <code>{user_key}</code>\n"
            f"🎫 Token    : {tok}  <i>(reset 00:00 WIB)</i>\n"
            f"📧 Email    : {email_count}/{email_limit} akun\n"
            f"💬 Grup     : {grup_status}\n\n"
            f"🛒 Upgrade paket → /beli"
            f"</blockquote>")
    else:
        sisa_hari = max(0, int((prem.get("expired", 0) - time.time()) // 86400))
        tok_str = "♾️ Unlimited" if t["tokens_day"] >= 99999 else f"{t['tokens_day']}/hari"
        send_msg(chat_id,
            f"📊 <b>STATUS AKUN</b>\n\n"
            f"<blockquote>"
            f"{t['emoji']} Paket    : <b>{t['label']}</b>\n"
            f"🔑 Key      : <code>{user_key}</code>\n"
            f"🎫 Token    : {tok}  <i>(reset 00:00 WIB)</i>\n"
            f"⚡ Limit    : {tok_str}\n"
            f"📅 Sisa     : {sisa_hari} hari\n"
            f"📧 Email    : {email_count}/{email_limit} akun\n"
            f"💬 Grup     : {grup_status}"
            f"</blockquote>")                        

# ================= MENU & COMMANDS SYSTEM =================
def handle_start(user_id, chat_id):
    owner  = is_owner(user_id)
    THUMBNAIL_PATH = "./thumbnail.png"
    tok = token_status_str(user_id)

    tier = get_user_tier(user_id)
    t_info = TOKEN_TIERS.get(tier, TOKEN_TIERS["free"])
    tier_badge = f"{t_info['emoji']} {t_info['label']}"

    if owner:
        caption = (
            "🤖 <b>BOT OTP IVAS V7</b>\n"
            "<i>SMS/OTP monitoring — Platform IVAS</i>\n\n"
            "👑 <b>OWNER PANEL</b>\n"
            "<blockquote>"
            "/addtoken — aktivasi paket user\n"
            "/deltoken — hapus paket user\n"
            "/listtoken — list paket aktif\n"
            "/setcookie\n"
            "/addakun\n"
            "/delakun\n"
            "/listakun\n"
            "/statsms"
            "</blockquote>\n\n"
            f"🎫 <b>Token:</b> {tok}\n\n"
            "🛠️ <b>FITUR</b>\n"
            "<blockquote>"
            "/addcookie\n"
            "/addemail email password\n"
            "/listemail\n"
            "/delcookie email\n"
            "/addnum\n"
            "/delnumall\n"
            "/myrange\n"
            "/ambilfile\n"
            "/cekivas\n"
            "/cekprem"
            "</blockquote>\n\n"
            "💬 <b>GROUP</b>\n"
            "<blockquote>"
            "/addgrup\n"
            "/delgrup\n"
            "/listgrup"
            "</blockquote>"
        )
    else:
        user_key = get_or_create_user_key(user_id)
        email_limit = get_tier_email_limit(user_id)
        caption = (
            "🤖 <b>BOT OTP IVAS V7</b>\n"
            "<i>SMS/OTP monitoring — Platform IVAS</i>\n\n"
            f"🔑 <b>Key:</b> <code>{user_key}</code>\n"
            f"🏷️ <b>Paket:</b> {tier_badge}\n"
            f"🎫 <b>Token:</b> {tok}  <i>(reset 00:00 WIB)</i>\n"
            f"📧 <b>Max Email:</b> {email_limit} akun\n\n"
            "🛠️ <b>FITUR</b> <i>(1 fitur = 1 token)</i>\n"
            "<blockquote>"
            "/addcookie\n"
            "/addemail email password\n"
            "/listemail\n"
            "/delcookie email\n"
            "/addnum\n"
            "/delnumall\n"
            "/myrange\n"
            "/ambilfile\n"
            "/cekivas\n"
            "/cekprem"
            "</blockquote>\n\n"
            "💬 <b>GROUP</b> <i>(gratis)</i>\n"
            "<blockquote>"
            "/addgrup\n"
            "/delgrup\n"
            "/listgrup"
            "</blockquote>\n\n"
            "🛒 <b>Upgrade Paket</b> → /beli\n"
            f"📩 <a href='https://t.me/kicenxensai'>@kicenxensai</a>"
        )

    try:
        if os.path.exists(THUMBNAIL_PATH):
            with open(THUMBNAIL_PATH, "rb") as photo:
                r = requests.post(
                    f"https://api.telegram.org/bot{BOT_TOKEN}/sendPhoto",
                    data={"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"},
                    files={"photo": photo},
                    timeout=15
                )
            if not r.json().get("ok"):
                send_msg(chat_id, caption)
        else:
            send_msg(chat_id, caption)
    except Exception as e:
        send_msg(chat_id, caption)

def code_to_flag(code):
    try: return ''.join(chr(127397 + ord(c)) for c in code.upper())
    except: return "  "
        
def add_email(text, chat_id, user_id, msg_id):
    try:
        parts = text.split()
        if len(parts) < 3:
            return send_msg(chat_id, "❌ Format:\n/addemail email@gmail.com password")
        email    = parts[1].strip().lower()
        password = parts[2].strip()
        if "@" not in email:
            return send_msg(chat_id, "❌ Email tidak valid!")

        users     = load_users()
        uid       = str(user_id)
        user_data = users.get(uid, {"emails": []})
        current_count = len(user_data.get("emails", []))

        # Limit email berdasarkan tier paket
        email_limit = get_tier_email_limit(user_id)
        if not is_owner(user_id) and current_count >= email_limit:
            tier = get_user_tier(user_id)
            t_d = TOKEN_TIERS.get(tier, TOKEN_TIERS["free"])
            if tier == "free":
                return send_msg(chat_id,
                    f"❌ <b>Limit akun FREE: {email_limit}</b>\n\n"
                    f"<blockquote>Upgrade paket untuk tambah lebih banyak akun.\n"
                    f"Ketik /beli untuk lihat paket tersedia.\n"
                    f"📩 <a href='https://t.me/kicenxensai'>@kicenxensai</a></blockquote>")
            else:
                return send_msg(chat_id,
                    f"❌ Limit paket {t_d['emoji']} <b>{t_d['label']}</b>: maksimal {email_limit} akun!\n\n"
                    f"<blockquote>Upgrade ke paket lebih tinggi via /beli</blockquote>")
        if email in user_data.get("emails", []):
            return send_msg(chat_id, "❌ Akun sudah ada!")

        # Simpan email ke list & simpan password di user_accounts
        user_data.setdefault("emails", []).append(email)
        user_accs = [a for a in user_data.get("user_accounts", []) if a.get("email") != email]
        user_accs.append({"email": email, "password": password})
        user_data["user_accounts"] = user_accs
        users[uid] = user_data
        save_users(users)

        # Coba login persis seperti add_account owner
        acc = {
            "email": email, "password": password,
            "cookies": {}, "session": make_httpx_client(),
            "last_login": 0, "csrf_token": ""
        }
        send_msg(chat_id, f"⏳ Mencoba login ke <code>{email}</code>...")
        if login(acc):
            acc["last_login"] = time.time()
            send_msg(chat_id, f"✅ <b>Akun aktif &amp; login:</b>\n<code>{email}</code>")
        else:
            send_msg(chat_id,
                f"⚠️ <b>Akun ditambahkan, tapi login gagal:</b>\n<code>{email}</code>\n\n"
                f"<blockquote>Coba /addcookie untuk pasang cookie manual.</blockquote>"
            )
    except Exception as e:
        send_msg(chat_id, f"❌ Error tambah akun: {e}")

def list_email(chat_id, user_id):
    users = load_users()
    if str(user_id) not in users or not users[str(user_id)]["emails"]: return send_msg(chat_id, "  Belum ada email")
    msg = "  <b>LIST EMAIL</b>\n\n"
    for i, em in enumerate(users[str(user_id)]["emails"], 1): msg += f"{i}. {em}\n"
    send_msg(chat_id, msg)        
        
def get_user_emails(user_id):
    """Kembalikan daftar email milik user: owner -> dari accounts, premium -> dari users.json"""
    if is_owner(user_id):
        return [acc["email"] for acc in accounts]
    users = load_users()
    return users.get(str(user_id), {}).get("emails", [])

# ================= ADDNUM FLOW =================
def command_addnum(text, chat_id, user_id):
    emails = get_user_emails(user_id)
    if not emails:
        return send_msg(chat_id, "❌ Belum ada email/akun.\nTambah dulu dengan /addemail atau /addakun")
    buttons = [{"text": f"📧 {em}", "callback_data": f"an:{em}"} for em in emails]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:an"})
    send_inline_keyboard(chat_id,
        "➕ <b>ADD NUMBER</b>\n\n"
        "<blockquote>📋 Cara Penggunaan:\n"
        "1. Pilih email akun IVAS di bawah\n"
        "2. Ketik target nomor atau negara\n"
        "   Contoh: <code>SAUDI ARABIA 15022</code>\n"
        "   Contoh: <code>INDONESIA 500</code>\n"
        "3. Bot akan proses penambahan nomor ke akun\n\n"
        "⚠️ Pastikan cookie sudah aktif sebelum add number</blockquote>\n\n"
        "👇 Pilih email:",
        buttons)

def handle_addnum_email_cb(chat_id, user_id, email, cb_id, msg_id):
    answer_callback_query(cb_id, "✅ Email dipilih!")
    emails = get_user_emails(user_id)
    if email not in emails:
        answer_callback_query(cb_id, "❌ Email tidak ditemukan")
        return
    new_msg_id = delete_and_send_with_cancel(chat_id, msg_id,
        f"➕ <b>ADD NUMBER</b>\n\n"
        f"📧 Email: <code>{email}</code>\n\n"
        f"<blockquote>✏️ Ketik range yang ingin ditambahkan:\n\n"
        f"<b>1 Range:</b>\n"
        f"<code>BENIN 851</code>\n\n"
        f"<b>Multi Range (pisah enter/koma):</b>\n"
        f"<code>BENIN 851\nMOZAMBIQUE 4234\nSAUDI ARABIA 15022</code></blockquote>",
        "an"
    )
    pending_addnum[user_id] = {"email": email, "msg_id": new_msg_id}

def _do_addnum_range(acc, session, csrf, target_text, progress_cb=None):
    """Fetch test numbers dan add untuk 1 range. Return dict hasil."""
    test_url = f"{BASE}/portal/numbers/test"
    hdrs = {
        "Accept":           "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Referer":          test_url,
    }

    def _fetch_test(length):
        p = {
            "draw":                   "1",
            "columns[0][data]":       "range",
            "columns[0][name]":       "terminations.range",
            "columns[1][data]":       "test_number",
            "columns[1][name]":       "terminations.test_number",
            "columns[2][data]":       "id",
            "columns[2][name]":       "id",
            "columns[3][data]":       "limit_did_a2p",
            "columns[3][name]":       "limit_did_a2p",
            "columns[4][data]":       "limit_cli_did_a2p",
            "columns[4][name]":       "limit_cli_did_a2p",
            "order[0][column]":       "0",
            "order[0][dir]":          "asc",
            "start":                  "0",
            "length":                 str(length),
            "search[value]":          target_text,
            "search[regex]":          "false",
        }
        r = session.get(test_url, params=p, headers=hdrs, timeout=20)
        if r.status_code != 200:
            raise Exception(f"HTTP {r.status_code}")
        return r.json()

    try:
        # Probe dulu untuk tahu total nomor yang tersedia (recordsFiltered)
        probe      = _fetch_test(1)
        total_avail = int(probe.get("recordsFiltered", probe.get("recordsTotal", 0)))
        # Gunakan total_avail agar semua nomor di range bisa di-fetch
        # Minimal fetch 100, maksimal 1000 agar tidak terlalu berat
        fetch_count = max(100, min(total_avail if total_avail > 0 else 1000, 1000))
        data = _fetch_test(fetch_count)
        rows = data.get("data", [])
    except Exception as e:
        return {"success": 0, "fail": 0, "skipped": False, "total": 0,
                "skip_msg": "", "not_found": False, "error": str(e)}

    fallback_fields = ["range", "test_number", "id", "limit_did_a2p", "limit_cli_did_a2p",
                       "term", "A2P", "created_at", "action"]
    rn_lower = target_text.lower().strip()
    items = []
    for row in rows:
        if isinstance(row, list):
            row = dict(zip(fallback_fields, row))
        rng = re.sub(r"<[^>]+>", "", str(row.get("range", ""))).strip()
        if rng.lower().strip() != rn_lower:
            continue
        tid = str(row.get("id", "") or row.get("DT_RowId", "")).strip()
        if tid and not tid.isdigit():
            m2 = re.search(r"(\d+)", tid)
            tid = m2.group(1) if m2 else ""
        if tid:
            items.append({"tid": tid})

    if not items:
        return {"success": 0, "fail": 0, "skipped": False, "total": 0,
                "skip_msg": "", "not_found": True, "error": None}

    add_url  = f"{BASE}/portal/numbers/termination/number/add"
    add_hdrs = {
        "Accept":           "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Referer":          f"{BASE}/portal/numbers/test",
        "Origin":           BASE,
        "Content-Type":     "application/x-www-form-urlencoded; charset=UTF-8",
    }

    success_count = 0
    fail_count    = 0
    skipped       = False
    skip_msg      = ""
    total_items   = len(items)
    _last_cb_at   = [0]  # track kapan terakhir kirim progress

    for idx, item in enumerate(items):
        tid = item["tid"]
        try:
            resp = session.post(add_url, data={"id": tid, "_token": csrf},
                                headers=add_hdrs, timeout=15)
            try:
                jr      = resp.json()
                message = str(jr.get("message", jr.get("msg", jr.get("error", str(jr)))))
                st      = jr.get("status", jr.get("success", ""))
                ok      = str(st).lower() in ("success", "ok", "true", "1") or st is True or st == 1
                if not ok:
                    ok = any(k in message.lower() for k in
                             ("berhasil", "success", "added", "good job", "successfully", "done"))
                if not ok and any(k in message.lower() for k in
                                  ("too many", "maximum", "limit", "penuh")):
                    skipped  = True
                    skip_msg = message
                    break
            except Exception:
                raw = resp.text.lower()
                ok  = any(k in raw for k in ("berhasil", "success", "added", "good job"))
                if any(k in raw for k in ("too many", "maximum", "limit", "penuh")):
                    skipped  = True
                    skip_msg = f"HTTP {resp.status_code}: limit tercapai"
                    break
            if ok:
                success_count += 1
            else:
                fail_count += 1
            time.sleep(0.25)
        except Exception:
            fail_count += 1

        # Kirim progress callback setiap 10 nomor atau di nomor terakhir
        done = idx + 1
        if progress_cb and (done - _last_cb_at[0] >= 10 or done == total_items):
            try:
                progress_cb(done, total_items, success_count, fail_count)
            except Exception:
                pass
            _last_cb_at[0] = done

    return {"success": success_count, "fail": fail_count, "skipped": skipped,
            "total": total_items, "skip_msg": skip_msg, "not_found": False, "error": None}


def process_addnum_target(chat_id, user_id, target_text):
    state = pending_addnum.pop(user_id, None)
    if not state:
        return False
    email  = state["email"]
    msg_id = state["msg_id"]

    # Parse multi-range: pisah per baris atau koma
    raw_ranges = re.split(r"[\n,]+", target_text)
    ranges = [r.strip() for r in raw_ranges if r.strip()]
    if not ranges:
        return False

    preview = ", ".join(f"<code>{r}</code>" for r in ranges[:3])
    if len(ranges) > 3:
        preview += f" +{len(ranges)-3} lainnya"

    proc_id = delete_and_send(chat_id, msg_id,
        f"➕ <b>ADD NUMBER</b>\n\n"
        f"<blockquote>"
        f"📧 Email: <code>{email}</code>\n"
        f"🎯 {'Range' if len(ranges) == 1 else f'{len(ranges)} Range'}: {preview}\n\n"
        f"⏳ Mencari nomor di range...</blockquote>")

    def _run():
        multi = len(ranges) > 1
        acc = None
        with accounts_lock:
            for a in accounts:
                if a.get("email") == email:
                    acc = a
                    break

        if not acc:
            delete_and_send(chat_id, proc_id,
                f"➕ <b>ADD NUMBER</b>\n\n"
                f"❌ Akun <code>{email}</code> tidak ditemukan.")
            return

        if not ensure_login(acc):
            delete_and_send(chat_id, proc_id,
                f"➕ <b>ADD NUMBER</b>\n\n"
                f"❌ Session akun <code>{email}</code> tidak aktif.\n"
                f"Gunakan /setcookie untuk memperbarui cookie.")
            return

        session = acc["session"]
        csrf    = acc.get("csrf_token", "")
        results = []

        for i, rng_target in enumerate(ranges):
            # Tampilkan status range saat ini (real-time)
            done_lines = ""
            for prev in results:
                if prev.get("error"):
                    st = "❌ Error"
                elif prev.get("not_found"):
                    st = "❌ Tdk ditemukan"
                elif prev["skipped"] and prev["success"] == 0:
                    st = "⚠️ Penuh"
                elif prev["success"] > 0:
                    st = f"✅ {prev['success']} nomor"
                else:
                    st = "❌ Gagal"
                done_lines += f"• <code>{prev['range']}</code>: {st}\n"

            if multi:
                edit_msg(chat_id, proc_id,
                    f"➕ <b>ADD NUMBER</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"⏳ [{i+1}/{len(ranges)}] Proses: <code>{rng_target}</code>...\n"
                    + (f"\n{done_lines.strip()}" if done_lines else "")
                    + f"</blockquote>")
            else:
                edit_msg(chat_id, proc_id,
                    f"➕ <b>ADD NUMBER</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"🎯 Range: <code>{rng_target}</code>\n\n"
                    f"⏳ Memulai add nomor...</blockquote>")

            # Progress callback — update pesan setiap 10 nomor (real-time)
            def make_progress_cb(rng_name, p_id, is_multi, i_idx, tot_ranges, d_lines):
                def _cb(done, total, ok, fail):
                    pct = int(done / total * 100) if total else 0
                    bar_filled = int(pct / 10)
                    bar = "▓" * bar_filled + "░" * (10 - bar_filled)
                    if is_multi:
                        edit_msg(chat_id, p_id,
                            f"➕ <b>ADD NUMBER</b>\n\n"
                            f"<blockquote>"
                            f"📧 Email: <code>{email}</code>\n"
                            f"⏳ [{i_idx+1}/{tot_ranges}] <code>{rng_name}</code>\n"
                            f"[{bar}] {pct}%\n"
                            f"✅ {ok} berhasil | ❌ {fail} gagal | 📊 {done}/{total}\n"
                            + (f"\n{d_lines.strip()}" if d_lines else "")
                            + f"</blockquote>")
                    else:
                        edit_msg(chat_id, p_id,
                            f"➕ <b>ADD NUMBER</b>\n\n"
                            f"<blockquote>"
                            f"📧 Email: <code>{email}</code>\n"
                            f"🎯 Range: <code>{rng_name}</code>\n\n"
                            f"[{bar}] {pct}%\n"
                            f"✅ {ok} berhasil | ❌ {fail} gagal | 📊 {done}/{total}</blockquote>")
                return _cb

            cb = make_progress_cb(rng_target, proc_id, multi, i, len(ranges), done_lines)
            r = _do_addnum_range(acc, session, csrf, rng_target, progress_cb=cb)
            results.append({"range": rng_target, **r})
            if i < len(ranges) - 1:
                time.sleep(0.5)

        if not multi:
            r = results[0]
            if r.get("error"):
                result_text = (
                    f"➕ <b>ADD NUMBER GAGAL</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"🎯 Range: <code>{ranges[0]}</code>\n"
                    f"❌ Gagal fetch IVAS: <code>{r['error'][:100]}</code>"
                    f"</blockquote>"
                )
            elif r.get("not_found"):
                result_text = (
                    f"➕ <b>ADD NUMBER GAGAL</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"🎯 Range: <code>{ranges[0]}</code>\n"
                    f"❌ Range tidak ditemukan di Test Numbers."
                    f"</blockquote>"
                )
            elif r["skipped"] and r["success"] == 0:
                result_text = (
                    f"➕ <b>ADD NUMBER GAGAL</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"🎯 Range: <code>{ranges[0]}</code>\n"
                    f"⚠️ Slot nomor di range ini sudah penuh\n\n"
                    f"Hubungi admin IVAS untuk tambah kuota."
                    f"</blockquote>"
                )
            elif r["skipped"]:
                result_text = (
                    f"➕ <b>ADD NUMBER SELESAI</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"✅ <code>{ranges[0]}</code>\n"
                    f"⚠️ Berhenti: Slot akun sudah penuh"
                    f"</blockquote>"
                )
            else:
                result_text = (
                    f"➕ <b>ADD NUMBER {'BERHASIL' if r['success'] > 0 else 'GAGAL'}</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"{'✅' if r['success'] > 0 else '❌'} <code>{ranges[0]}</code>"
                    f"</blockquote>"
                )
        else:
            total_ok   = sum(1 for r in results if r.get("success", 0) > 0)
            total_fail = sum(1 for r in results if r.get("success", 0) == 0 and not r.get("skipped") and not r.get("error") and not r.get("not_found"))
            lines = ""
            for r in results:
                if r.get("error"):
                    status = "❌ Error fetch"
                elif r.get("not_found"):
                    status = "❌ Tidak ditemukan"
                elif r["skipped"] and r["success"] == 0:
                    status = "⚠️ Penuh"
                elif r["skipped"]:
                    status = "✅ (lalu penuh)"
                elif r["success"] > 0:
                    status = "✅"
                else:
                    status = "❌ Gagal"
                lines += f"• <code>{r['range']}</code>: {status}\n"

            result_text = (
                f"➕ <b>ADD NUMBER SELESAI</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"🔢 Total: ✅ <b>{total_ok}</b> berhasil | ❌ <b>{total_fail}</b> gagal\n\n"
                f"{lines.strip()}"
                f"</blockquote>"
            )

        if multi:
            edit_msg(chat_id, proc_id, result_text)
        else:
            delete_and_send(chat_id, proc_id, result_text)

    threading.Thread(target=_run, daemon=True).start()
    return True


# ================= DELNUMALL FLOW =================
def command_delnumall(text, chat_id, user_id):
    emails = get_user_emails(user_id)
    if not emails:
        return send_msg(chat_id, "❌ Belum ada email/akun.")
    buttons = [{"text": f"📧 {em}", "callback_data": f"da:{em}"} for em in emails]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:da"})
    send_inline_keyboard(chat_id,
        "🗑️ <b>DELETE ALL NUMBER</b>\n\n"
        "<blockquote>📋 Cara Penggunaan:\n"
        "1. Pilih email akun IVAS di bawah\n"
        "2. Bot akan otomatis return semua nomor yang aktif\n"
        "3. Tunggu konfirmasi selesai\n\n"
        "⚠️ Semua nomor akan dikembalikan ke pool IVAS!</blockquote>\n\n"
        "👇 Pilih email:",
        buttons)

def handle_delnumall_email_cb(chat_id, user_id, email, cb_id, msg_id):
    answer_callback_query(cb_id, "⏳ Memproses...")
    emails = get_user_emails(user_id)
    if email not in emails:
        delete_and_send(chat_id, msg_id,
            "🗑️ <b>DELETE ALL NUMBER</b>\n\n❌ Email tidak ditemukan.")
        return
    proc_id = delete_and_send(chat_id, msg_id,
        f"🗑️ <b>DELETE ALL NUMBER</b>\n\n"
        f"<blockquote>"
        f"📧 Email: <code>{email}</code>\n\n"
        f"⏳ Sedang menghapus semua nomor..."
        f"</blockquote>")

    acc_target = next((a for a in accounts if a["email"] == email), None)
    if not acc_target:
        prem_cookies = load_premium_cookies()
        if email not in prem_cookies:
            delete_and_send(chat_id, proc_id,
                f"🗑️ <b>DELETE ALL NUMBER</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Akun/cookie tidak ditemukan. Set cookie dulu."
                f"</blockquote>")
            return
        session = make_httpx_client()
        session.cookies.update(prem_cookies[email])
        acc_target = {"email": email, "session": session, "last_login": time.time(),
                      "password": "", "csrf_token": "", "cookies": prem_cookies[email]}

    if not ensure_login(acc_target):
        delete_and_send(chat_id, proc_id,
            f"🗑️ <b>DELETE ALL NUMBER</b>\n\n"
            f"<blockquote>"
            f"📧 Email: <code>{email}</code>\n"
            f"❌ Gagal login/verifikasi session. Perbarui cookie."
            f"</blockquote>")
        return

    ok, res = return_all_base(acc_target)
    if ok:
        delete_and_send(chat_id, proc_id,
            f"🗑️ <b>DELETE ALL NUMBER BERHASIL</b>\n\n"
            f"<blockquote>"
            f"📧 Email: <code>{email}</code>\n"
            f"✅ Semua nomor berhasil dikembalikan ke pool!"
            f"</blockquote>")
    else:
        delete_and_send(chat_id, proc_id,
            f"🗑️ <b>DELETE ALL NUMBER GAGAL</b>\n\n"
            f"<blockquote>"
            f"📧 Email: <code>{email}</code>\n"
            f"❌ {str(res)[:150]}"
            f"</blockquote>")

# ================= MYRANGE FLOW =================
def command_myrange(text, chat_id, user_id):
    emails = get_user_emails(user_id)
    if not emails:
        return send_msg(chat_id, "❌ Belum ada email/akun.")
    buttons = [{"text": f"📧 {em}", "callback_data": f"mr:{em}"} for em in emails]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:mr"})
    send_inline_keyboard(chat_id,
        "📊 <b>MY RANGE</b>\n\n"
        "<blockquote>📋 Cara Penggunaan:\n"
        "1. Pilih email akun IVAS di bawah\n"
        "2. Bot akan menampilkan semua range di My Numbers\n"
        "3. Termasuk jumlah nomor per range</blockquote>\n\n"
        "👇 Pilih email:",
        buttons)

def handle_myrange_email_cb(chat_id, user_id, email, cb_id, msg_id):
    answer_callback_query(cb_id, "⏳ Memproses...")
    emails = get_user_emails(user_id)
    if email not in emails:
        delete_and_send(chat_id, msg_id,
            "📊 <b>MY RANGE</b>\n\n❌ Email tidak ditemukan.")
        return

    proc_id = delete_and_send(chat_id, msg_id,
        f"📊 <b>MY RANGE</b>\n\n"
        f"<blockquote>"
        f"📧 Email: <code>{email}</code>\n\n"
        f"⏳ Sedang mengambil data range..."
        f"</blockquote>")

    acc_target = next((a for a in accounts if a["email"] == email), None)
    if not acc_target:
        prem_cookies = load_premium_cookies()
        if email not in prem_cookies:
            delete_and_send(chat_id, proc_id,
                f"📊 <b>MY RANGE</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Akun/cookie tidak ditemukan. Set cookie dulu."
                f"</blockquote>")
            return
        session = make_httpx_client()
        session.cookies.update(prem_cookies[email])
        acc_target = {"email": email, "session": session, "last_login": time.time(),
                      "password": "", "csrf_token": "", "cookies": prem_cookies[email]}

    if not ensure_login(acc_target):
        delete_and_send(chat_id, proc_id,
            f"📊 <b>MY RANGE</b>\n\n"
            f"<blockquote>"
            f"📧 Email: <code>{email}</code>\n"
            f"❌ Gagal login/verifikasi session. Perbarui cookie."
            f"</blockquote>")
        return

    try:
        my_url = f"{BASE}/portal/numbers"
        # Kolom CONFIRMED dari file referensi — harus 8 kolom
        col_data = ["Number", "range", "A2P", "LimitA2P", "limit_did_a2p", "limit_cli_a2p", "number_id", "action"]
        col_name = ["Number", "range", "A2P",  "LimitA2P", "limit_did_a2p", "limit_cli_a2p", "number_id", "action"]
        col_qs = "".join(
            f"&columns[{i}][data]={d}&columns[{i}][name]={n}"
            for i, (d, n) in enumerate(zip(col_data, col_name))
        )
        qs = (
            f"draw=1{col_qs}"
            "&order[0][column]=0&order[0][dir]=asc"
            "&start=0&length=2000"
            "&search[value]=&search[regex]=false"
        )
        hdrs = {
            "Accept":           "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest",
            "Referer":          my_url,
        }
        session = acc_target["session"]
        resp = session.get(f"{my_url}?{qs}", headers=hdrs, timeout=20)
        if resp.status_code != 200:
            raise Exception(f"HTTP {resp.status_code}")
        data = resp.json()
        rows = data.get("data", [])
        total = data.get("recordsTotal", 0)

        # Handle rows sebagai list-of-lists (convert ke dict)
        if rows and isinstance(rows[0], list):
            rows = [dict(zip(col_data, r)) for r in rows]

        from collections import Counter
        range_count = Counter()
        for row in rows:
            if isinstance(row, dict):
                rng = re.sub(r"<[^>]+>", "", str(row.get("range", ""))).strip()
                if rng:
                    range_count[rng] += 1

        if not range_count:
            delete_and_send(chat_id, proc_id,
                f"📊 <b>MY RANGE</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"ℹ️ Tidak ada nomor di My Numbers."
                f"</blockquote>")
            return

        lines = ""
        for i, (rng, cnt) in enumerate(sorted(range_count.items()), 1):
            lines += f"{i}. <b>{rng}</b> — {cnt} nomor\n"

        result_text = (
            f"📊 <b>MY RANGE</b>\n\n"
            f"<blockquote>"
            f"📧 Email: <code>{email}</code>\n"
            f"🔢 Total: <b>{total}</b> nomor | <b>{len(range_count)}</b> range\n\n"
            f"{lines.strip()}"
            f"</blockquote>"
        )
        delete_and_send(chat_id, proc_id, result_text)

    except Exception as ex:
        delete_and_send(chat_id, proc_id,
            f"📊 <b>MY RANGE</b>\n\n"
            f"<blockquote>"
            f"📧 Email: <code>{email}</code>\n"
            f"❌ Error: {str(ex)[:150]}"
            f"</blockquote>")

# ================= AMBILFILE FLOW =================
def command_ambilfile(text, chat_id, user_id):
    emails = get_user_emails(user_id)
    if not emails:
        return send_msg(chat_id, "❌ Belum ada email/akun.")
    buttons = [{"text": f"📧 {em}", "callback_data": f"af:{em}"} for em in emails]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:af"})
    send_inline_keyboard(chat_id,
        "📁 <b>AMBIL FILE</b>\n\n"
        "<blockquote>📋 Cara Penggunaan:\n"
        "1. Pilih email akun IVAS di bawah\n"
        "2. Bot akan mengambil data nomor dari IVAS\n"
        "3. File Excel (.xlsx) dikirim otomatis ke chat ini\n\n"
        "💡 File berisi semua nomor aktif beserta range/negara</blockquote>\n\n"
        "👇 Pilih email:",
        buttons)

def handle_ambilfile_email_cb(chat_id, user_id, email, cb_id, msg_id):
    answer_callback_query(cb_id, "⏳ Memproses...")
    emails = get_user_emails(user_id)
    if email not in emails:
        delete_and_send(chat_id, msg_id,
            "📁 <b>AMBIL FILE</b>\n\n❌ Email tidak ditemukan.")
        return

    proc_id = delete_and_send(chat_id, msg_id,
        f"📁 <b>AMBIL FILE</b>\n\n"
        f"<blockquote>"
        f"📧 Email: <code>{email}</code>\n\n"
        f"⏳ Sedang mengambil &amp; menyusun file export..."
        f"</blockquote>")

    all_cookies = load_cookies()
    prem_cookies = load_premium_cookies()
    cookies = all_cookies.get(email) or prem_cookies.get(email)
    if not cookies:
        delete_and_send(chat_id, proc_id,
            f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
            f"<blockquote>"
            f"📧 Email: <code>{email}</code>\n"
            f"❌ Cookie tidak ditemukan. Set cookie dulu."
            f"</blockquote>")
        return

    acc_target = {"email": email, "cookies": cookies, "csrf_token": ""}
    export_numbers_ivas(chat_id, acc_target, status_msg_id=proc_id)

def delete_msg(chat_id, message_id):
    try: requests.post(f"https://api.telegram.org/bot{BOT_TOKEN}/deleteMessage", data={"chat_id": chat_id, "message_id": message_id}, timeout=10)
    except: pass    

def detect_country_and_flag(full_num, fallback_country="UNKNOWN"):
    try:
        parsed = phonenumbers.parse("+" + full_num, None)
        region = phonenumbers.region_code_for_number(parsed)
        if region:
            flag = code_to_flag(region)
            country_name = geocoder.description_for_number(parsed, "en")
            if not country_name: country_name = fallback_country
            return country_name.upper(), flag
    except Exception as e: print("FLAG ERROR:", e)
    return fallback_country, "  "
    
def parse_cookie_input(raw_text):
    try:
        data = json.loads(raw_text)
        if isinstance(data, list):
            cookie_dict = {}
            for item in data:
                if isinstance(item, dict) and "name" in item and "value" in item:
                    cookie_dict[item["name"]] = item["value"]
            return cookie_dict if cookie_dict else None
        elif isinstance(data, dict):
            return data
        return None
    except:
        return None

def get_recv_csrf(acc) -> str:
    """
    Ambil CSRF token dari halaman /portal/sms/received.
    iVAS pakai per-page rotating CSRF — semua POST ke getsms API
    WAJIB pakai token dari halaman ini, bukan dari /portal umum.
    Di-cache 15 menit per akun.
    """
    email = acc.get("email", "")
    now   = time.time()
    cached = _recv_csrf_cache.get(email)
    if cached and (now - cached["ts"]) < RECV_CSRF_TTL:
        return cached["csrf"]
    try:
        r = acc["session"].get(RECV_URL, timeout=15)
        if "/login" in str(r.url):
            return acc.get("recv_csrf") or acc.get("csrf_token", "")
        soup = BeautifulSoup(r.text, "html.parser")
        csrf = ""
        meta = soup.find("meta", {"name": "csrf-token"})
        if meta:
            csrf = meta.get("content", "")
        if not csrf:
            inp = soup.find("input", {"name": "_token"})
            if inp:
                csrf = inp.get("value", "")
        if not csrf:
            m = re.search(r"['\"]_token['\"]\s*[,:]?\s*['\"]([A-Za-z0-9_\-+/=]{20,})['\"]", r.text)
            if m:
                csrf = m.group(1)
        if csrf:
            acc["recv_csrf"] = csrf
            _recv_csrf_cache[email] = {"csrf": csrf, "ts": now}
            return csrf
    except Exception as e:
        print(f"WARN get_recv_csrf [{email}]: {e}")
    return acc.get("recv_csrf") or acc.get("csrf_token", "")


def verify_cookie_session(acc):
    """
    Verifikasi session cookie.
    1. GET /portal — cek tidak redirect ke /login & ambil csrf_token umum
    2. GET /portal/sms/received — ambil recv_csrf khusus untuk SMS API
    """
    try:
        session = acc["session"]
        r = session.get(f"{BASE}/portal", timeout=15)
        if "/login" in str(r.url):
            return False
        soup = BeautifulSoup(r.text, "html.parser")
        token_input = soup.find("input", {"name": "_token"})
        if token_input:
            acc["csrf_token"] = token_input["value"]
        else:
            token_meta = soup.find("meta", {"name": "csrf-token"})
            if token_meta:
                acc["csrf_token"] = token_meta.get("content", "")
        # Ambil recv_csrf — WAJIB untuk POST ke SMS API
        email = acc.get("email", "")
        _recv_csrf_cache.pop(email, None)  # Paksa refresh recv_csrf setelah verify
        get_recv_csrf(acc)
        return True
    except Exception as e:
        print(f"Cookie verify error: {e}")
        return False

def send_inline_keyboard(chat_id, text, buttons):
    keyboard = {"inline_keyboard": [[{"text": b["text"], "callback_data": b["callback_data"]}] for b in buttons]}
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML", "reply_markup": keyboard},
            timeout=10
        )
        return r.json().get("result", {}).get("message_id")
    except:
        return None

def send_inline_keyboard_grid(chat_id, text, rows):
    """Kirim pesan dengan inline keyboard grid (2D array). rows = [[btn, btn], [btn], ...]"""
    def make_btn(b):
        if "url" in b:
            return {"text": b["text"], "url": b["url"]}
        return {"text": b["text"], "callback_data": b["callback_data"]}
    keyboard = {"inline_keyboard": [[make_btn(b) for b in row] for row in rows]}
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML", "reply_markup": keyboard},
            timeout=10
        )
        return r.json().get("result", {}).get("message_id")
    except:
        return None

def edit_msg(chat_id, message_id, text, remove_keyboard=False):
    try:
        payload = {"chat_id": chat_id, "message_id": message_id, "text": text, "parse_mode": "HTML"}
        if remove_keyboard:
            payload["reply_markup"] = {"inline_keyboard": []}
        requests.post(f"https://api.telegram.org/bot{BOT_TOKEN}/editMessageText", json=payload, timeout=10)
    except:
        pass

def delete_and_send(chat_id, msg_id, text):
    """Hapus pesan lama, kirim pesan baru. Return message_id baru."""
    delete_msg(chat_id, msg_id)
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
            timeout=10
        )
        return r.json().get("result", {}).get("message_id")
    except:
        return None

def delete_and_send_keyboard(chat_id, msg_id, text, buttons):
    """Hapus pesan lama, kirim pesan baru dengan inline keyboard. Return message_id baru."""
    delete_msg(chat_id, msg_id)
    return send_inline_keyboard(chat_id, text, buttons)

def send_with_cancel(chat_id, text, cancel_key):
    """Kirim pesan baru dengan tombol ❌ Batalkan. Return message_id."""
    keyboard = {"inline_keyboard": [[{"text": "❌ Batalkan", "callback_data": f"cancel:{cancel_key}"}]]}
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML", "reply_markup": keyboard},
            timeout=10
        )
        return r.json().get("result", {}).get("message_id")
    except:
        return None

def delete_and_send_with_cancel(chat_id, msg_id, text, cancel_key):
    """Hapus pesan lama, kirim pesan baru dengan tombol ❌ Batalkan. Return message_id baru."""
    delete_msg(chat_id, msg_id)
    return send_with_cancel(chat_id, text, cancel_key)

def answer_callback_query(callback_query_id, text=""):
    requests.post(
        f"https://api.telegram.org/bot{BOT_TOKEN}/answerCallbackQuery",
        data={"callback_query_id": callback_query_id, "text": text}
    )

# ================= SETCOOKIE FLOW (OWNER) =================
def cmd_setcookie(chat_id):
    if not accounts:
        send_msg(chat_id, "❌ Belum ada akun. Tambah dulu dengan /addakun")
        return
    buttons = [{"text": f"📧 {acc['email']}", "callback_data": f"setcookie:{acc['email']}"} for acc in accounts]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:sc"})
    send_inline_keyboard(chat_id,
        "🍪 <b>SET COOKIE — OWNER</b>\n\n"
        "<blockquote>📋 Cara Penggunaan:\n"
        "1. Pilih email akun IVAS di bawah\n"
        "2. Kirim full JSON cookie dari browser\n"
        "3. Bot akan verifikasi session otomatis\n\n"
        "💡 Export cookie: DevTools → Application → Cookies</blockquote>\n\n"
        "👇 Pilih email:",
        buttons
    )

def handle_setcookie_callback(chat_id, user_id, email, callback_query_id, msg_id):
    answer_callback_query(callback_query_id, "✅ Email dipilih!")
    new_msg_id = delete_and_send_with_cancel(chat_id, msg_id,
        f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
        f"📧 Email: <code>{email}</code>\n\n"
        f"<blockquote>📤 Sekarang kirim full JSON cookie kamu.\n\n"
        f"Format array (export browser):\n"
        f"<code>[{{\"name\":\"key\",\"value\":\"val\"}}]</code>\n\n"
        f"Atau format dict:\n"
        f"<code>{{\"laravel_session\":\"...\",\"XSRF-TOKEN\":\"...\"}}</code></blockquote>",
        "sc"
    )
    pending_setcookie[user_id] = {"email": email, "msg_id": new_msg_id}

def process_cookie_input(chat_id, user_id, text):
    state = pending_setcookie.pop(user_id, None)
    if not state:
        return False

    email = state["email"]
    msg_id = state["msg_id"]

    cookie_dict = parse_cookie_input(text)
    if not cookie_dict:
        new_id = delete_and_send_with_cancel(chat_id, msg_id,
            f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
            f"📧 Email: <code>{email}</code>\n\n"
            f"❌ <b>Format JSON tidak valid!</b>\n"
            f"<blockquote>Kirim ulang cookie dalam format yang benar.</blockquote>",
            "sc"
        )
        pending_setcookie[user_id] = {"email": email, "msg_id": new_id}
        return True

    proc_id = delete_and_send(chat_id, msg_id,
        f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
        f"📧 Email: <code>{email}</code>\n\n"
        f"⏳ Memverifikasi cookie..."
    )

    found = False
    with accounts_lock:
        for acc in accounts:
            if acc["email"] == email:
                found = True
                if "session" not in acc or acc["session"] is None:
                    acc["session"] = make_httpx_client()
                acc["session"].cookies.clear()
                acc["session"].cookies.update(cookie_dict)

                # Reset session fail flags agar run_bot tidak skip akun ini
                _session_notified[email] = False
                _session_fail_time.pop(email, None)
                _session_retry_time.pop(email, None)
                _session_recovered.pop(email, None)
                acc["last_login"] = 0

                if verify_cookie_session(acc):
                    acc["last_login"] = time.time()
                    # Ambil fresh cookies setelah verifikasi berhasil
                    fresh = extract_session_cookies(acc["session"])
                    cookies_to_save = fresh if fresh else cookie_dict
                    acc["cookies"] = cookies_to_save
                    all_cdata = load_cookies()
                    all_cdata[email] = cookies_to_save
                    save_cookies(all_cdata)
                    # Reset keepalive timer & session fail flags — langsung aktif tanpa restart
                    _last_cookie_refresh[email] = time.time()
                    _session_notified[email] = False
                    _session_fail_time.pop(email, None)
                    _session_retry_time.pop(email, None)
                    _session_recovered.pop(email, None)
                    # Hapus ranges cache — paksa fetch fresh saat poll berikutnya
                    _ranges_cache.pop(email, None)
                    delete_and_send(chat_id, proc_id,
                        f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
                        f"✅ <b>Cookie berhasil disimpan &amp; langsung aktif!</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email: <code>{email}</code>\n"
                        f"🔑 Total cookie: <b>{len(cookies_to_save)}</b> key\n"
                        f"✔️ Session aktif &amp; terverifikasi\n"
                        f"🔄 Fresh cookie langsung dipakai (tanpa restart)"
                        f"</blockquote>"
                    )
                else:
                    # Cookie tidak valid — jangan simpan, beri pesan ramah
                    acc["session"].cookies.clear()
                    if acc.get("cookies"):
                        acc["session"].cookies.update(acc["cookies"])
                    delete_and_send(chat_id, proc_id,
                        f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
                        f"❌ <b>Cookie tidak valid / expired!</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email: <code>{email}</code>\n"
                        f"🔑 Total cookie dikirim: <b>{len(cookie_dict)}</b> key\n\n"
                        f"Cookie ini tidak bisa login ke server IVAS.\n"
                        f"Silakan ambil cookie <b>fresh</b> dari browser dan coba lagi 😊\n\n"
                        f"💡 Tips: Buka DevTools → Application → Cookies → copy semua"
                        f"</blockquote>"
                    )
                return True

    if not found:
        delete_and_send(chat_id, proc_id,
            f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
            f"❌ Email <code>{email}</code> tidak ditemukan di daftar akun."
        )
    return True

# ================= ADDCOOKIE FLOW (TOKEN) =================
def verify_cookie_dict(cookie_dict):
    try:
        session = make_httpx_client(timeout=15)
        session.cookies.update(cookie_dict)
        r = session.get(f"{BASE}/portal", timeout=15)
        return "/login" not in str(r.url)
    except:
        return False

def cmd_addcookie(chat_id, user_id):
    users = load_users()
    emails = users.get(str(user_id), {}).get("emails", [])
    if not emails:
        send_msg(chat_id, "❌ Belum ada email. Tambah dulu dengan /addemail")
        return
    buttons = [{"text": f"📧 {em}", "callback_data": f"addcookie:{em}"} for em in emails]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:ac"})
    send_inline_keyboard(chat_id,
        "🍪 <b>ADD COOKIE — TOKEN</b>\n\n"
        "<blockquote>📋 Cara Penggunaan:\n"
        "1. Pilih email IVAS kamu di bawah\n"
        "2. Kirim full JSON cookie dari browser\n"
        "3. Bot akan verifikasi session otomatis\n\n"
        "💡 Export cookie: DevTools → Application → Cookies</blockquote>\n\n"
        "👇 Pilih email:",
        buttons
    )

def handle_addcookie_callback(chat_id, user_id, email, callback_query_id, msg_id):
    answer_callback_query(callback_query_id, "✅ Email dipilih!")
    users = load_users()
    emails = users.get(str(user_id), {}).get("emails", [])
    if email not in emails:
        answer_callback_query(callback_query_id, "❌ Email tidak ditemukan")
        return
    new_msg_id = delete_and_send_with_cancel(chat_id, msg_id,
        f"🍪 <b>ADD COOKIE — TOKEN</b>\n\n"
        f"📧 Email: <code>{email}</code>\n\n"
        f"<blockquote>📤 Sekarang kirim full JSON cookie kamu.\n\n"
        f"Format array (export browser):\n"
        f"<code>[{{\"name\":\"key\",\"value\":\"val\"}}]</code>\n\n"
        f"Atau format dict:\n"
        f"<code>{{\"laravel_session\":\"...\",\"XSRF-TOKEN\":\"...\"}}</code></blockquote>",
        "ac"
    )
    pending_addcookie[user_id] = {"email": email, "msg_id": new_msg_id}

def process_addcookie_input(chat_id, user_id, text):
    state = pending_addcookie.pop(user_id, None)
    if not state:
        return False

    email = state["email"]
    msg_id = state["msg_id"]

    cookie_dict = parse_cookie_input(text)
    if not cookie_dict:
        new_id = delete_and_send_with_cancel(chat_id, msg_id,
            f"🍪 <b>ADD COOKIE — TOKEN</b>\n\n"
            f"📧 Email: <code>{email}</code>\n\n"
            f"❌ <b>Format JSON tidak valid!</b>\n"
            f"<blockquote>Kirim ulang cookie dalam format yang benar.</blockquote>",
            "ac"
        )
        pending_addcookie[user_id] = {"email": email, "msg_id": new_id}
        return True

    proc_id = delete_and_send(chat_id, msg_id,
        f"🍪 <b>ADD COOKIE — TOKEN</b>\n\n"
        f"📧 Email: <code>{email}</code>\n\n"
        f"⏳ Menyimpan &amp; memverifikasi cookie..."
    )

    try:
        # Verifikasi dulu pakai session sementara — jangan simpan sebelum verified
        tmp_session = make_httpx_client(timeout=15)
        tmp_session.cookies.update(cookie_dict)
        tmp_acc = {"session": tmp_session, "email": email, "cookies": cookie_dict, "csrf_token": ""}
        valid = verify_cookie_session(tmp_acc)

        if valid:
            # Ambil fresh cookies dari session setelah verifikasi berhasil
            fresh = extract_session_cookies(tmp_session)
            cookies_to_save = fresh if fresh else cookie_dict

            prem_cookies = load_premium_cookies()
            prem_cookies[email] = cookies_to_save
            save_premium_cookies(prem_cookies)

            # Langsung update _premium_acc_cache tanpa tunggu sync 30 detik
            if email in _premium_acc_cache:
                cached = _premium_acc_cache[email]
                cached["cookies"] = cookies_to_save
                cached["csrf_token"] = tmp_acc.get("csrf_token", "")
                cached["last_login"] = time.time()
                cached["session"].cookies.clear()
                cached["session"].cookies.update(cookies_to_save)
            else:
                new_acc = {
                    "email": email, "password": None,
                    "cookies": cookies_to_save,
                    "session": make_httpx_client(),
                    "last_login": time.time(),
                    "csrf_token": tmp_acc.get("csrf_token", ""),
                }
                new_acc["session"].cookies.update(cookies_to_save)
                _premium_acc_cache[email] = new_acc

            # Reset keepalive timer agar diprioritaskan saat ping berikutnya
            _last_cookie_refresh[email] = time.time()
            # Reset session fail flags
            _session_notified[email] = False
            _session_fail_time.pop(email, None)
            _session_retry_time.pop(email, None)
            _session_recovered.pop(email, None)
            # Hapus ranges cache — paksa fetch fresh saat poll berikutnya
            _ranges_cache.pop(email, None)
            # Paksa run_bot sync segera agar thread worker baru langsung spawn
            global _force_bot_sync
            _force_bot_sync = True

            delete_and_send(chat_id, proc_id,
                f"🍪 <b>ADD COOKIE — TOKEN</b>\n\n"
                f"✅ <b>Cookie berhasil disimpan &amp; langsung aktif!</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"🔑 Total cookie: <b>{len(cookies_to_save)}</b> key\n"
                f"✔️ Session aktif &amp; terverifikasi\n"
                f"🔄 Fresh cookie langsung dipakai (tanpa restart)"
                f"</blockquote>"
            )
        else:
            # Cookie tidak valid — jangan simpan, beri pesan ramah
            delete_and_send(chat_id, proc_id,
                f"🍪 <b>ADD COOKIE — TOKEN</b>\n\n"
                f"❌ <b>Cookie tidak valid / expired!</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"🔑 Total cookie dikirim: <b>{len(cookie_dict)}</b> key\n\n"
                f"Cookie ini tidak bisa login ke server IVAS.\n"
                f"Silakan ambil cookie <b>fresh</b> dari browser dan coba lagi 😊\n\n"
                f"💡 Tips: Buka DevTools → Application → Cookies → copy semua"
                f"</blockquote>"
            )
    except Exception as e:
        delete_and_send(chat_id, proc_id,
            f"🍪 <b>ADD COOKIE — TOKEN</b>\n\n"
            f"❌ Terjadi error saat verifikasi: <code>{e}</code>"
        )
    return True

def ensure_login(acc):
    now = time.time()
    email = acc.get("email", "")
    if now - acc.get("last_login", 0) < LOGIN_COOLDOWN:
        return True

    if acc.get("cookies"):
        if "session" not in acc or acc["session"] is None:
            acc["session"] = make_httpx_client()
        acc["session"].cookies.clear()
        acc["session"].cookies.update(acc["cookies"])
        if verify_cookie_session(acc):
            acc["last_login"] = now
            fresh = extract_session_cookies(acc["session"])
            if fresh and fresh != acc.get("cookies"):
                acc["cookies"] = fresh
                save_fresh_cookies_auto(email, fresh)
            if _session_notified.get(email):
                _session_notified[email] = False
                _session_fail_time.pop(email, None)
                _session_retry_time.pop(email, None)
                if not _session_recovered.get(email):
                    _session_recovered[email] = True
                    requests.post(
                        f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                        data={"chat_id": OWNER_ID, "text": f"✅ <b>SESSION PULIH</b>\n\n📧 Email: <code>{email}</code>\nSession berhasil aktif kembali secara otomatis.", "parse_mode": "HTML"},
                        timeout=10
                    )
            return True
        print(Fore.YELLOW + f"  COOKIE EXPIRED [{email}] — coba login password")

    if login(acc):
        acc["last_login"] = now
        if _session_notified.get(email):
            _session_notified[email] = False
            _session_fail_time.pop(email, None)
            _session_retry_time.pop(email, None)
            if not _session_recovered.get(email):
                _session_recovered[email] = True
                requests.post(
                    f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                    data={"chat_id": OWNER_ID, "text": f"✅ <b>SESSION PULIH</b>\n\n📧 Email: <code>{email}</code>\nLogin password berhasil, session aktif kembali.", "parse_mode": "HTML"},
                    timeout=10
                )
        return True

    if not _session_notified.get(email):
        _session_notified[email] = True
        _session_recovered[email] = False
        _session_fail_time[email] = now
        print(Fore.RED + f"  SESSION GAGAL [{email}] — notif dikirim")
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data={
                "chat_id": OWNER_ID,
                "text": (
                    f"⚠️ <b>SESSION EXPIRED</b>\n\n"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Cookie expired & login password gagal.\n\n"
                    f"Bot akan otomatis retry setiap 10 menit.\n"
                    f"Perbarui cookie dengan /setcookie atau /addcookie."
                ),
                "parse_mode": "HTML"
            },
            timeout=10
        )

    _session_retry_time[email] = now
    return False
 
def cek_ivas(chat_id=None):
    try:
        url = "http://ws.websocket.web.id/api/cekivas?platform=whatsapp"
        r = requests.get(url, timeout=10)
        send_to = chat_id if chat_id else OWNER_ID
        if r.status_code != 200: return send_msg(send_to, "  Gagal ambil data IVAS")
        data = r.json()
        if not data.get("success"): return send_msg(send_to, "  API gagal")
        results = data.get("results", [])
        if not results: return send_msg(send_to, "   Tidak ada data IVAS")

        results = sorted(results, key=lambda x: x["count"], reverse=True)
        msg = "  <b>CEK IVAS WHATSAPP</b>\n\n"
        for i, item in enumerate(results, 1):
            msg += f"{i}. {item.get('country', 'Unknown').upper()} : {item.get('count', 0)} SMS\n"
        send_msg(send_to, msg)
    except Exception as e:
        send_to = chat_id if chat_id else OWNER_ID
        send_msg(send_to, f"  Error cek IVAS: {e}")

# ================= UTILS =================
def extract_otp(text):
    m = re.search(r"\b(\d{3}[- ]?\d{3})\b", text)
    if not m: return None
    otp = m.group(0).replace(" ", "")  
    if len(otp) not in (6, 7): return None
    if len(otp) == 6: otp = otp[:3] + "-" + otp[3:]
    return otp    
        
def return_all_base(acc):
    try:
        session = acc["session"]
        url = RETURN_ALL_URL
        headers = {"X-Requested-With": "XMLHttpRequest", "Referer": f"{BASE}/portal/numbers", "Origin": BASE}
        r = session.post(url, headers=headers, data={"_token": acc.get("csrf_token", "")})
        if r.status_code == 200: return True, r.text
        else: return False, f"HTTP {r.status_code}"
    except Exception as e:
        return False, str(e)
        
def parse_range(rng):
    country = re.sub(r"\s*\(.*?\)", "", rng)
    country = re.sub(r"\d+", "", country)
    country = re.sub(r"\s+", " ", country).strip().upper()
    code_match = re.search(r"\((\d+)\)", rng)
    code = code_match.group(1) if code_match else ""
    return country, code

def extract_service_short(text):
    m = re.search(r"(WhatsApp|Telegram|Google|Facebook|Instagram|Shopee|Tokopedia|Grab|Gojek|TikTok)", text, re.I)
    if m: return SERVICE_SHORT.get(m.group(1).upper(), "#OT")
    return "#OT"

def mask_email(email):
    try:
        name, domain = email.split("@")
        if len(name) <= 2: return name + "*" + "@" + domain
        return f"{name[0]}{'*' * (len(name)-2)}{name[-1]}@{domain}"
    except:
        return email

def stats_sms(chat_id=None):
    total_sms = sms_stats["total_sms"]
    total_otp = sms_stats["total_otp"]
    total_number = len(sms_stats["total_number"])
    msg = f"  <b>STATISTIK SMS OTP</b>\n\n  Total SMS Masuk : {total_sms}\n  Total OTP       : {total_otp}\n  Total Nomor     : {total_number}\n  Total Akun Aktif: {len(accounts)}\n"
    if chat_id:
        send_msg(chat_id, msg)
    else:
        tg_active(msg)                        

def login(acc):
    session = acc["session"]
    email = acc["email"]
    password = acc["password"]

    r = session.get(LOGIN_URL)
    soup = BeautifulSoup(r.text, "html.parser")
    token_input = soup.find("input", {"name": "_token"})

    if not token_input:
        print("  CSRF TOKEN TIDAK DITEMUKAN")
        return False

    token = token_input["value"]
    acc["csrf_token"] = token

    session.headers.update({
        "X-CSRF-TOKEN": token,
        "X-Requested-With": "XMLHttpRequest"
    })

    r2 = session.post(LOGIN_URL, data={
        "_token": token,
        "email": email,
        "password": password
    })

    print("LOGIN RESPONSE URL:", r2.url)

    if "/portal" in str(r2.url) or "Dashboard" in r2.text or "portal" in r2.text.lower():
        print("  LOGIN BERHASIL")
        fresh = extract_session_cookies(session)
        if fresh:
            acc["cookies"] = fresh
            all_cookies = load_cookies()
            all_cookies[email] = fresh
            save_cookies(all_cookies)
            print(f"  FRESH COOKIE SAVED: {email} ({len(fresh)} keys)")
        return True
    else:
        print("  LOGIN GAGAL")
        return False

def _is_login_page(r) -> bool:
    """Cek apakah response adalah redirect ke halaman login (session expired)."""
    try:
        if "/login" in str(r.url):
            return True
        if r.status_code in (401, 403, 419):
            return True
    except Exception:
        pass
    return False


def _invalidate_session(acc, reason="SESSION_EXPIRED"):
    """
    Force re-verify session pada iterasi berikutnya.
    Hapus ranges cache & recv_csrf cache agar semua di-fetch ulang setelah re-login.
    """
    email = acc.get("email", "")
    acc["last_login"] = 0
    if email:
        _ranges_cache.pop(email, None)
        _recv_csrf_cache.pop(email, None)
    raise Exception(reason)


_RECV_POST_HEADERS = {
    "Accept":           "text/html,*/*;q=0.01",
    "X-Requested-With": "XMLHttpRequest",
    "Content-Type":     "application/x-www-form-urlencoded; charset=UTF-8",
    "Referer":          RECV_URL,  # WAJIB — iVAS CSRF middleware cek Referer
    "Origin":           BASE,
}


def get_ranges(acc):
    today = datetime.now().strftime("%Y-%m-%d")
    csrf  = get_recv_csrf(acc)
    r = acc["session"].post(GET_RANGE_URL,
        data={"_token": csrf, "from": today, "to": today},
        headers=_RECV_POST_HEADERS
    )
    if _is_login_page(r):
        _invalidate_session(acc, f"SESSION_EXPIRED: get_ranges ({r.url})")
    soup = BeautifulSoup(r.text, "html.parser")
    ranges = []
    for div in soup.find_all("div", onclick=True):
        if "toggleRange" in div["onclick"]:
            try: ranges.append(div["onclick"].split("'")[1])
            except: pass
    return list(set(ranges))

def get_ranges_cached(acc):
    """Cache ranges 5 menit. Auto-invalidate saat session expired."""
    email = acc.get("email", "")
    now   = time.time()
    entry = _ranges_cache.get(email)
    if entry:
        ts, cached_ranges = entry
        if now - ts < RANGES_CACHE_TTL:
            return cached_ranges
    ranges = get_ranges(acc)
    _ranges_cache[email] = (now, ranges)
    return ranges

def get_numbers(acc, rng):
    today = datetime.now().strftime("%Y-%m-%d")
    csrf  = get_recv_csrf(acc)
    r = acc["session"].post(GET_NUMBER_URL,
        data={"_token": csrf, "start": today, "end": today, "range": rng},
        headers=_RECV_POST_HEADERS
    )
    if _is_login_page(r):
        _invalidate_session(acc, f"SESSION_EXPIRED: get_numbers ({r.url})")
    soup = BeautifulSoup(r.text, "html.parser")
    numbers = []
    for div in soup.find_all("div", onclick=True):
        try:
            val = div["onclick"].split("'")[1]
            if val and val != rng: numbers.append(val)
        except: pass
    return list(set(numbers))

def get_sms(acc, rng, number):  
    today = datetime.now().strftime("%Y-%m-%d")
    csrf  = get_recv_csrf(acc)
    r = acc["session"].post(GET_SMS_URL,
        data={"_token": csrf, "start": today, "end": today, "Number": number, "Range": rng},
        headers=_RECV_POST_HEADERS
    )
    if _is_login_page(r):
        _invalidate_session(acc, f"SESSION_EXPIRED: get_sms ({r.url})")
    soup = BeautifulSoup(r.text, "html.parser")  
    sms_texts = []  
    try:  
        texts = list(soup.stripped_strings)  
        for t in texts:  
            t = t.strip()  
            if t.startswith("<#>"): t = t.replace("<#>", "").strip()  
            if re.fullmatch(r"[A-Za-z0-9]{10,}", t): continue  
            t_low = t.lower()  
            if any(x in t_low for x in ["sender", "revenue", "time"]): continue  
            if re.search(r"\b\d{2}:\d{2}:\d{2}\b", t): continue  
            if "$" in t: continue  
            if t and "No SMS Found" not in t: sms_texts.append(t)  
    except Exception as e: print("ERROR PARSE SMS:", e)  
    return list(dict.fromkeys(sms_texts))  
    
def format_phone_number(number):
    number = str(number).replace("+", "").replace(" ", "")
    if len(number) >= 10:
        return f"{number[:4]}****{number[-4:]}"
    return number    
    
def normalize_number(num, country_code):
    num = str(num).strip().replace(" ", "").replace("-", "").replace("+", "")
    if num.startswith(country_code): return num
    if num.startswith("0"): return country_code + num[1:]
    return num

def tg_active(msg):
    _tg_request("sendMessage", data={"chat_id": OWNER_ID, "text": msg, "parse_mode": "HTML"})
            
# ================= TELEGRAM LISTENER =================
_TG_POLL_CLIENT = httpx.Client(
    follow_redirects=True,
    timeout=35,
    headers={"User-Agent": "Mozilla/5.0"},
)

def listen_command():
    global last_update_id
    _backoff = 0
    while True:
        try:
            url = f"https://api.telegram.org/bot{BOT_TOKEN}/getUpdates"
            r = _TG_POLL_CLIENT.get(url, params={"offset": last_update_id + 1, "timeout": 25})
            data = r.json()

            for upd in data.get("result", []):
                last_update_id = upd["update_id"]

                # ====== HANDLE CALLBACK QUERY (inline button click) ======
                if "callback_query" in upd:
                    try:
                        cq = upd["callback_query"]
                        cq_id = cq["id"]
                        cq_data = cq.get("data", "")
                        cq_user_id = cq["from"]["id"]
                        cq_chat_id = cq["message"]["chat"]["id"]
                        cq_msg_id = cq["message"]["message_id"]

                        if cq_data == "check_join":
                            not_joined = check_force_join(cq_user_id)
                            if not_joined:
                                answer_callback_query(cq_id, "⚠️ Kamu belum join semua channel/grup!")
                                send_force_join_msg(cq_chat_id, not_joined)
                            else:
                                answer_callback_query(cq_id, "✅ Sudah join semua! Silakan gunakan bot.")
                                delete_msg(cq_chat_id, cq_msg_id)
                                handle_start(cq_user_id, cq_chat_id)
                        elif cq_data.startswith("setcookie:"):
                            if is_owner(cq_user_id):
                                handle_setcookie_callback(cq_chat_id, cq_user_id, cq_data[len("setcookie:"):], cq_id, cq_msg_id)
                            else:
                                answer_callback_query(cq_id, "❌ Khusus OWNER")
                        elif cq_data.startswith("pkg_info:"):
                            handle_pkg_info_cb(cq_chat_id, cq_user_id, cq_data[9:], cq_id, cq_msg_id)
                        elif cq_data.startswith("pkg_buy:"):
                            threading.Thread(target=handle_pkg_buy_cb, args=(cq_chat_id, cq_user_id, cq_data[8:], cq_id, cq_msg_id), daemon=True).start()
                        elif cq_data == "pkg_back":
                            answer_callback_query(cq_id)
                            delete_msg(cq_chat_id, cq_msg_id)
                            cmd_beli(cq_chat_id, cq_user_id)
                        elif cq_data.startswith("addcookie:"):
                            handle_addcookie_callback(cq_chat_id, cq_user_id, cq_data[len("addcookie:"):], cq_id, cq_msg_id)
                        elif cq_data.startswith("an:"):
                            handle_addnum_email_cb(cq_chat_id, cq_user_id, cq_data[3:], cq_id, cq_msg_id)
                        elif cq_data.startswith("da:"):
                            threading.Thread(target=handle_delnumall_email_cb, args=(cq_chat_id, cq_user_id, cq_data[3:], cq_id, cq_msg_id), daemon=True).start()
                        elif cq_data.startswith("af:"):
                            threading.Thread(target=handle_ambilfile_email_cb, args=(cq_chat_id, cq_user_id, cq_data[3:], cq_id, cq_msg_id), daemon=True).start()
                        elif cq_data.startswith("mr:"):
                            threading.Thread(target=handle_myrange_email_cb, args=(cq_chat_id, cq_user_id, cq_data[3:], cq_id, cq_msg_id), daemon=True).start()
                        elif cq_data.startswith("cancel_payment:"):
                            parts = cq_data[len("cancel_payment:"):].split(":")
                            if len(parts) == 2:
                                cp_order_id, cp_amount_str = parts
                                pending = pending_payments.get(cq_user_id, {})
                                if pending.get("order_id") == cp_order_id:
                                    pending_payments.pop(cq_user_id, None)
                                    answer_callback_query(cq_id, "❌ Pembayaran dibatalkan")
                                    delete_msg(cq_chat_id, cq_msg_id)
                                    threading.Thread(target=pakasir_cancel, args=(cp_order_id, int(cp_amount_str)), daemon=True).start()
                                    send_msg(cq_chat_id,
                                        "❌ <b>Pembayaran dibatalkan.</b>\n\n"
                                        "Tagihan QRIS sudah dibatalkan.\n"
                                        "Ketik /beli untuk membuat tagihan baru."
                                    )
                                else:
                                    answer_callback_query(cq_id, "Tidak ada tagihan aktif")
                            else:
                                answer_callback_query(cq_id)
                        elif cq_data.startswith("cancel:"):
                            answer_callback_query(cq_id, "❌ Dibatalkan")
                            key = cq_data[7:]
                            if key == "sc":
                                pending_setcookie.pop(cq_user_id, None)
                            elif key == "ac":
                                pending_addcookie.pop(cq_user_id, None)
                            elif key == "an":
                                pending_addnum.pop(cq_user_id, None)
                            delete_msg(cq_chat_id, cq_msg_id)
                            send_msg(cq_chat_id, "❌ <b>Aksi dibatalkan.</b>")
                        else:
                            answer_callback_query(cq_id)
                    except Exception as ex:
                        print(f"Error callback_query: {ex}")
                    continue

                if "message" not in upd: continue
                try:
                    msg = upd["message"]
                    text = msg.get("text", "") or ""
                    user_id = msg["from"]["id"]
                    chat_id = msg["chat"]["id"]
                    msg_id = msg["message_id"]

                    store_username(user_id, msg["from"])

                    owner = is_owner(user_id)
                    is_group = msg["chat"]["type"] in ["group", "supergroup"]

                    # ====== CEK WAJIB JOIN (hanya user non-owner di private chat) ======
                    if not owner and not is_group and text.startswith("/") and text != "/start":
                        not_joined = check_force_join(user_id)
                        if not_joined:
                            send_force_join_msg(chat_id, not_joined)
                            continue

                    # ====== CEK PENDING SETCOOKIE (owner input cookie JSON) ======
                    if owner and user_id in pending_setcookie and text and not text.startswith("/"):
                        if process_cookie_input(chat_id, user_id, text):
                            continue

                    # ====== CEK PENDING ADDCOOKIE (semua user bisa, input cookie JSON) ======
                    if user_id in pending_addcookie and text and not text.startswith("/"):
                        if process_addcookie_input(chat_id, user_id, text):
                            continue

                    # ====== CEK PENDING ADDNUM (semua user bisa, asal sedang dalam pending state) ======
                    if user_id in pending_addnum and text and not text.startswith("/"):
                        if process_addnum_target(chat_id, user_id, text):
                            continue

                    # ROUTING COMMAND TEXT
                    udisp = get_user_display(user_id)

                    if text == "/start":
                        if not owner:
                            not_joined = check_force_join(user_id)
                            if not_joined:
                                send_force_join_msg(chat_id, not_joined)
                                continue
                        handle_start(user_id, chat_id)
                    elif text.startswith("/cekivas"):
                        if use_token(user_id):
                            cek_ivas(chat_id)
                            send_activity_log(user_id, udisp, "/cekivas")
                        else: no_token_msg(chat_id)
                    elif text.startswith("/cekprem"): cek_premium(chat_id, user_id)
                    
                    elif text.startswith("/listakun"): 
                        if owner: list_accounts(chat_id, user_id)
                        else: send_msg(chat_id, "  Khusus OWNER")
                    elif text.startswith("/addcookie"):
                        if use_token(user_id):
                            add_cookie_premium(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/addcookie")
                        else: no_token_msg(chat_id)
                    elif text.startswith("/delcookie"):
                        if use_token(user_id):
                            del_cookie_premium(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/delcookie")
                        else: no_token_msg(chat_id)
                    
                    elif text.startswith("/addemail"):
                        if use_token(user_id):
                            add_email(text, chat_id, user_id, msg_id)
                            send_activity_log(user_id, udisp, "/addemail")
                        else: no_token_msg(chat_id)
                    elif text.startswith("/listemail"): list_email(chat_id, user_id)
                    
                    elif text.startswith("/addgrup"):
                        if is_group:
                            gid = str(chat_id)
                            if gid in get_user_groups(user_id):
                                send_msg(chat_id, "  Grup sudah ada di akun kamu")
                                send_activity_log(user_id, udisp, "/addgrup", "⚠️ Grup sudah ada")
                            else:
                                add_user_group(user_id, gid)
                                send_msg(chat_id, f"✅ <b>Grup berhasil ditambahkan!</b>\n\n<blockquote>🆔 ID: <code>{gid}</code>\n🔑 Key: <code>{get_or_create_user_key(user_id)}</code></blockquote>")
                                send_activity_log(user_id, udisp, "/addgrup")
                        else: send_msg(chat_id, "  Jalankan di dalam grup!")

                    elif text.startswith("/delgrup"):
                        gid = str(chat_id)
                        if remove_user_group(user_id, gid):
                            send_msg(chat_id, f"  Grup dihapus dari akun kamu:\n{gid}")
                            send_activity_log(user_id, udisp, "/delgrup")
                        else: send_msg(chat_id, "  Grup tidak ditemukan di akun kamu")

                    elif text.startswith("/listgrup"):
                        my_groups = get_user_groups(user_id)
                        if not my_groups: send_msg(chat_id, "Belum ada grup di akun kamu")
                        else:
                            msg_out = "  <b>LIST GRUP KAMU</b>\n\n"
                            for i, g in enumerate(my_groups, 1): msg_out += f"{i}. <code>{g}</code>\n"
                            send_msg(chat_id, msg_out)

                    elif text.startswith("/addnum"):
                        if use_token(user_id):
                            command_addnum(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/addnum")
                        else: no_token_msg(chat_id)
                        
                    elif text.startswith("/ambilfile"):
                        if use_token(user_id):
                            command_ambilfile(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/ambilfile")
                        else: no_token_msg(chat_id)

                    elif text.startswith("/delnumall"):
                        if use_token(user_id):
                            command_delnumall(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/delnumall")
                        else: no_token_msg(chat_id)

                    elif text.startswith("/myrange"):
                        if use_token(user_id):
                            command_myrange(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/myrange")
                        else: no_token_msg(chat_id)
                    
                    elif text.startswith("/beli"):
                        cmd_beli(chat_id, user_id)
                        send_activity_log(user_id, udisp, "/beli", "📦 Buka menu pembelian")
                    elif text.startswith("/addtoken"): 
                        if owner: add_token_tier(text, chat_id) 
                        else: send_msg(chat_id, "❌ Khusus OWNER")
                    elif text.startswith("/deltoken"): 
                        if owner: del_token_tier(text, chat_id) 
                        else: send_msg(chat_id, "❌ Khusus OWNER")
                    elif text.startswith("/listtoken"): 
                        if owner: list_token_tier(chat_id) 
                        else: send_msg(chat_id, "❌ Khusus OWNER")
                    elif text.startswith("/addakun"): 
                        if owner: add_account(text) 
                        else: send_msg(chat_id, "  Khusus OWNER")
                    elif text.startswith("/delakun"): 
                        if owner: del_account(text) 
                        else: send_msg(chat_id, "  Khusus OWNER")
                    elif text.startswith("/setcookie"): 
                        if owner: cmd_setcookie(chat_id)
                        else: send_msg(chat_id, "  Khusus OWNER")
                    elif text.startswith("/statsms"): 
                        if owner: stats_sms(chat_id) 
                        else: send_msg(chat_id, "  Khusus OWNER")
                except Exception as ex: 
                    print(f"Error handling message: {ex}")
        except Exception as e:
            _backoff = min(_backoff + 2, 20)
            print(f"Loop listener error: {e} — retry in {_backoff}s")
            time.sleep(_backoff)

            
# ================= POLL ENGINE (per-account) =================

def poll_one_account(acc):
    """
    Satu iterasi polling SMS untuk satu akun.
    Return True jika ada SMS/OTP baru ditemukan, False jika tidak.
    """
    email = acc.get("email", "")
    if not email:
        return False

    found_sms = False  # flag — diset True jika ada OTP dikirim

    # Cek session retry
    if _session_notified.get(email):
        if time.time() - _session_retry_time.get(email, 0) < SESSION_RETRY_INTERVAL:
            return False
        print(f"  AUTO-RETRY SESSION: {email}")
        acc["last_login"] = 0

    if not ensure_login(acc):
        return False

    owner_uid  = _bot_state["email_to_uid"].get(email, OWNER_ID)
    total      = _bot_state["total_accounts"]

    # ── Tentukan tujuan kirim: hanya ke pemilik akun (user/grup mereka sendiri) ──
    acct_groups  = get_user_groups(owner_uid)
    send_targets = acct_groups if acct_groups else [str(owner_uid)]

    try:
        ranges = get_ranges_cached(acc)
    except Exception as e:
        err = str(e)
        if "SESSION_EXPIRED" not in err:
            print(Fore.YELLOW + f"  WARN get_ranges [{email}]: {err}")
        return False

    for rng in ranges:
        fallback_country, code = parse_range(rng)
        try:
            numbers = get_numbers(acc, rng)
        except Exception as e:
            err = str(e)
            if "SESSION_EXPIRED" not in err:
                print(Fore.YELLOW + f"  WARN get_numbers [{email}]: {err}")
            continue

        for num in numbers:
            full_num = normalize_number(num, code)
            if not full_num.isdigit():
                continue

            try:
                sms_list = get_sms(acc, rng, num)
            except Exception as e:
                print(f"WARN get_sms [{email}]: {e}")
                continue

            for sms in sms_list:
                clean_sms = re.sub(r"\s+", " ", sms.replace("<#>", "")).strip()
                sms_uid = hashlib.md5(f"{num}-{clean_sms}".encode()).hexdigest()

                with _sent_cache_lock:
                    if sms_uid in sent_cache:
                        continue

                matches = re.findall(r"\b\d{3}[- ]?\d{3}\b", sms)
                if not matches:
                    continue

                otp = matches[0].replace(" ", "-")
                masked_num = format_phone_number(full_num)
                service_name = extract_service_short(sms)
                country, flag = detect_country_and_flag(full_num, fallback_country)
                clean_sms_display = clean_sms[:300]

                msg = (
                    f"<b>🔔 OTP BARU DITERIMA!</b>\n\n"
                    f"<blockquote><b>📱 Nomor :</b> <code>{masked_num}</code></blockquote>\n"
                    f"<b>🔑 OTP :</b> <code>{otp}</code>\n"
                    f"<blockquote><b>🛒 Service :</b> {service_name}</blockquote>\n"
                    f"<blockquote><b>🌍 Negara :</b> {country} {flag}</blockquote>\n"
                    f"<blockquote><b>📧 Email :</b> {mask_email(email)}</blockquote>\n"
                    f"<blockquote><b>📊 Total Aktif :</b> {total} Akun</blockquote>\n\n"
                    f"💬 <b>Pesan:</b>\n"
                    f"<code>{clean_sms_display}</code>"
                )

                for gid in send_targets:
                    res = _tg_request("sendMessage",
                                data={"chat_id": gid, "text": msg, "parse_mode": "HTML"},
                                timeout=10)
                    if res is not None and not res.json().get("ok"):
                        err = res.json().get("description", "unknown error")
                        print(Fore.RED + f"  SEND GAGAL → {gid}: {err}")

                with _sent_cache_lock:
                    sent_cache.add(sms_uid)
                save_sent_cache_debounced()

                sms_stats["total_sms"] += 1
                sms_stats["total_otp"] += 1
                if len(sms_stats["total_number"]) < 10000:
                    sms_stats["total_number"].add(full_num)

                user_display = get_user_display(owner_uid)
                print(Fore.GREEN + f"OTP → {user_display} ({mask_email(email)}) | {masked_num} | {otp}")
                found_sms = True

    return found_sms


# ================= AUTO COOKIE REFRESHER =================
def _notify_cookie_expired(email, uid):
    """Kirim notif ke pemilik akun bahwa cookie expired — max 1x per COOKIE_NOTIF_COOLDOWN."""
    now = time.time()
    if now - _last_cookie_notif.get(email, 0) < COOKIE_NOTIF_COOLDOWN:
        return
    _last_cookie_notif[email] = now

    msg = (
        f"⚠️ <b>COOKIE EXPIRED — AUTO REFRESH GAGAL</b>\n\n"
        f"📧 Email: <code>{email}</code>\n"
        f"❌ Cookie sudah expired dan tidak bisa auto-login.\n\n"
        f"<blockquote>Silakan perbarui cookie dengan:\n"
        f"• Owner  : /setcookie\n"
        f"• User   : /addcookie\n\n"
        f"💡 Ambil cookie fresh dari browser:\n"
        f"DevTools → Application → Cookies → copy semua</blockquote>"
    )
    # Kirim notif HANYA ke pemilik akun (bukan bocor ke owner kalau akun milik user)
    target = uid if uid else OWNER_ID
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data={"chat_id": target, "text": msg, "parse_mode": "HTML"},
            timeout=10
        )
    except Exception as e:
        print(Fore.RED + f"  NOTIF ERROR [{email}]: {e}")


def auto_cookie_refresher():
    """
    Background keepalive — ping server tiap 10 menit per akun.

    Cara kerja (tanpa password):
    1. Hit endpoint portal dengan session + cookie yang ada
    2. Server Laravel otomatis extend session lifetime
    3. Extract fresh cookies dari response → simpan ke file
    4. Session tidak pernah sempat expired selama bot hidup

    Jika session sudah benar-benar expired (bot restart lama / downtime):
    → Notif owner + user untuk update cookie manual, TIDAK coba password login
    """
    print(Fore.CYAN + "  AUTO KEEPALIVE — background aktif (ping tiap 10 menit)")
    time.sleep(90)  # Tunggu bot fully ready + semua akun ter-load dulu

    while True:
        try:
            now = time.time()

            # Kumpulkan semua akun (owner + premium)
            with accounts_lock:
                owner_accs = list(accounts)
            prem_accs = list(_premium_acc_cache.values())
            all_accs = owner_accs + prem_accs

            # Filter hanya akun yang sudah waktunya di-ping
            due = [
                a for a in all_accs
                if a.get("email") and
                   a.get("cookies") and
                   now - _last_cookie_refresh.get(a["email"], 0) >= COOKIE_KEEPALIVE_INTERVAL
            ]

            if due:
                pass  # Silent ping — log hanya muncul jika ada error/warning

            for acc in due:
                email = acc["email"]
                try:
                    # Gunakan session utama akun — bukan throwaway session
                    # Supaya session yang sama dipakai bot terus terjaga fresh
                    if "session" not in acc or acc["session"] is None:
                        acc["session"] = make_httpx_client()

                    # Pasang cookies terkini ke session utama sebelum ping
                    stored = acc.get("cookies", {})
                    if stored:
                        acc["session"].cookies.clear()
                        acc["session"].cookies.update(stored)

                    # Ping ke portal — server Laravel extend session pada setiap request
                    r = acc["session"].get(f"{BASE}/portal", timeout=15)

                    if r.status_code == 200 and "/login" not in str(r.url):
                        # Session masih hidup — ambil fresh cookies + update CSRF
                        soup = BeautifulSoup(r.text, "html.parser")
                        t = soup.find("input", {"name": "_token"})
                        if t:
                            acc["csrf_token"] = t["value"]

                        fresh = extract_session_cookies(acc["session"])
                        if fresh:
                            acc["cookies"] = fresh
                            acc["last_login"] = now
                            save_fresh_cookies_auto(email, fresh)
                            # Refresh recv_csrf — hapus cache lama agar GET /portal/sms/received dipanggil ulang
                            _recv_csrf_cache.pop(email, None)
                            get_recv_csrf(acc)
                            # Reset flag session gagal jika sebelumnya sempat error
                            if _session_notified.get(email):
                                _session_notified[email] = False
                                _session_fail_time.pop(email, None)
                                _session_retry_time.pop(email, None)
                                if not _session_recovered.get(email):
                                    _session_recovered[email] = True
                                    _uid_recover = _bot_state.get("email_to_uid", {}).get(email, OWNER_ID)
                                    recover_target = _uid_recover if _uid_recover else OWNER_ID
                                    try:
                                        requests.post(
                                            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                                            data={
                                                "chat_id": recover_target,
                                                "text": (
                                                    f"✅ <b>SESSION PULIH</b>\n\n"
                                                    f"📧 Email: <code>{email}</code>\n"
                                                    f"Session aktif kembali secara otomatis."
                                                ),
                                                "parse_mode": "HTML"
                                            },
                                            timeout=10
                                        )
                                    except Exception:
                                        pass
                            _keepalive_warn_count[email] = 0
                            print(Fore.GREEN + f"  KEEPALIVE OK: {email} — {len(fresh)} cookie di-extend")
                    else:
                        # Hitung berapa kali keepalive gagal berturut-turut
                        fail_n = _keepalive_warn_count.get(email, 0) + 1
                        _keepalive_warn_count[email] = fail_n
                        uid = _bot_state.get("email_to_uid", {}).get(email, OWNER_ID)

                        if fail_n == 1:
                            # Kegagalan PERTAMA → kirim warning awal (sebelum konfirmasi expired)
                            print(Fore.YELLOW + f"  KEEPALIVE WARNING ({fail_n}x): {email}")
                            last_notif = _last_cookie_notif.get(email + "_warn", 0)
                            if now - last_notif > COOKIE_NOTIF_COOLDOWN:
                                _last_cookie_notif[email + "_warn"] = now
                                warn_msg = (
                                    f"⚠️ <b>SESSION WARNING</b>\n\n"
                                    f"📧 Email: <code>{email}</code>\n"
                                    f"Session tidak merespons. Kemungkinan cookie akan segera expired.\n\n"
                                    f"<blockquote>Bot sedang otomatis retry...\n"
                                    f"Jika berlanjut, segera perbarui cookie dengan:\n"
                                    f"• Owner: /setcookie\n"
                                    f"• User: /addcookie</blockquote>"
                                )
                                # Kirim HANYA ke pemilik akun (tidak bocor ke owner jika akun milik user)
                                warn_target = uid if uid else OWNER_ID
                                try:
                                    requests.post(
                                        f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                                        data={"chat_id": warn_target, "text": warn_msg, "parse_mode": "HTML"},
                                        timeout=10
                                    )
                                except Exception:
                                    pass
                        else:
                            # Kegagalan ke-2+ → session benar-benar expired, notif hard
                            print(Fore.RED + f"  KEEPALIVE EXPIRED ({fail_n}x): {email} — session mati, notif user")
                            _notify_cookie_expired(email, uid)
                            if not _session_notified.get(email):
                                _session_notified[email] = True
                                _session_recovered[email] = False
                                _session_fail_time[email] = now

                    _last_cookie_refresh[email] = now

                except Exception as e:
                    print(Fore.RED + f"  KEEPALIVE ERROR [{email}]: {e}")

                time.sleep(2)  # Jeda kecil antar akun — jangan hammering server

            time.sleep(60)  # Loop setiap 1 menit untuk cek akun mana yang due

        except Exception as e:
            print(Fore.RED + f"ERROR AUTO KEEPALIVE: {e}")
            time.sleep(60)


# ================= AUTO BACKUP =================
def _collect_backup_files():
    """
    Kumpulkan semua file project secara rekursif, skip folder/file sistem.
    Return list of (absolute_path, archive_name).
    """
    root = os.path.abspath(".")
    collected = []
    for dirpath, dirnames, filenames in os.walk(root):
        # Pruning — hapus dir yang harus diskip dari traversal
        dirnames[:] = [
            d for d in dirnames
            if d not in BACKUP_SKIP_DIRS and not d.startswith(".")
            or d in {"file", "voice"}          # folder project yang harus masuk
        ]
        # Tambahan: pastikan folder tersembunyi non-project tetap diskip
        dirnames[:] = [
            d for d in dirnames
            if not (d.startswith(".") and d not in {"file", "voice"})
            and d not in BACKUP_SKIP_DIRS
        ]

        for fname in filenames:
            # Skip berdasarkan ekstensi
            _, ext = os.path.splitext(fname)
            if ext.lower() in BACKUP_SKIP_EXTS:
                continue
            # Skip file tertentu
            if fname in BACKUP_SKIP_FILES:
                continue
            # Skip file .zip di root (hasil backup lama)
            abs_path = os.path.join(dirpath, fname)
            rel_path = os.path.relpath(abs_path, root)
            collected.append((abs_path, rel_path))

    return collected


def _send_backup_telegram():
    """Scan seluruh project, buat ZIP, dan kirim ke owner Telegram."""
    now_str  = datetime.now().strftime("%d-%m-%Y_%H%M")
    zip_name = f"BACKUPAN SC IVAS KICEN_{now_str}.zip"
    # Simpan zip di luar root agar tidak ikut ter-scan
    zip_path = f"/tmp/{zip_name}"
    try:
        files_to_backup = _collect_backup_files()
        total_files = len(files_to_backup)

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for abs_path, arc_name in files_to_backup:
                try:
                    zf.write(abs_path, arc_name)
                except Exception:
                    pass  # skip file yang tidak bisa dibaca

        size_kb = round(os.path.getsize(zip_path) / 1024, 1)
        size_mb = round(size_kb / 1024, 2)
        size_str = f"{size_mb} MB" if size_kb > 1024 else f"{size_kb} KB"

        with open(zip_path, "rb") as f:
            requests.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                data={
                    "chat_id": OWNER_ID,
                    "caption": (
                        f"📦 <b>AUTO BACKUP — {datetime.now().strftime('%d %b %Y %H:%M')}</b>\n\n"
                        f"🗂️ File: <code>{zip_name}</code>\n"
                        f"📁 Total: <b>{total_files} file</b> (termasuk semua folder)\n"
                        f"📏 Ukuran: <b>{size_str}</b>"
                    ),
                    "parse_mode": "HTML",
                },
                files={"document": (zip_name, f, "application/zip")},
                timeout=120,
            )
        print(Fore.GREEN + f"  AUTO BACKUP TERKIRIM: {zip_name} ({size_str}, {total_files} file)")
    except Exception as e:
        print(Fore.RED + f"  AUTO BACKUP ERROR: {e}")
        try:
            requests.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                data={
                    "chat_id": OWNER_ID,
                    "text": f"❌ <b>Auto Backup Gagal</b>\n\n<code>{e}</code>",
                    "parse_mode": "HTML",
                },
                timeout=10,
            )
        except Exception:
            pass
    finally:
        try:
            if os.path.exists(zip_path):
                os.remove(zip_path)
        except Exception:
            pass


def run_auto_backup():
    """Background thread: backup langsung saat startup, lalu tiap jam 00:00."""
    print(Fore.CYAN + "  AUTO BACKUP — background aktif (startup + setiap jam 00:00)")
    time.sleep(20)  # Tunggu bot & akun selesai init

    # ── Backup pertama: langsung saat startup ─────────────────────────────────
    print(Fore.CYAN + "  AUTO BACKUP STARTUP — kirim backup awal...")
    _send_backup_telegram()

    # ── Loop: backup berikutnya setiap tengah malam ───────────────────────────
    while True:
        try:
            now = datetime.now()
            next_midnight = (now + timedelta(days=1)).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            wait_sec = (next_midnight - now).total_seconds()

            jam   = int(wait_sec // 3600)
            menit = int((wait_sec % 3600) // 60)
            print(Fore.CYAN + f"  BACKUP berikutnya dalam {jam}j {menit}m ({next_midnight.strftime('%d %b %Y 00:00')})")

            time.sleep(wait_sec)

            print(Fore.CYAN + "  AUTO BACKUP TENGAH MALAM — dimulai...")
            _send_backup_telegram()

            time.sleep(65)  # Jeda agar tidak trigger 2x di menit yang sama

        except Exception as e:
            print(Fore.RED + f"ERROR AUTO BACKUP: {e}")
            time.sleep(3600)


# ================= EXPIRY NOTIFIER (paket premium + cookie) =================
def _send_expiry_notif(uid, tier, sisa_detik, level):
    """Kirim notifikasi paket akan expired ke user dan owner."""
    t_info     = TOKEN_TIERS.get(tier, {})
    tier_label = t_info.get("label", tier.upper())
    tier_emoji = t_info.get("emoji", "🏷️")

    jam   = int(sisa_detik // 3600)
    menit = int((sisa_detik % 3600) // 60)
    sisa_str = f"{jam} jam {menit} menit" if jam > 0 else f"{menit} menit"

    if level == "24h":
        icon, judul = "⚠️", "PAKET AKAN EXPIRED — 24 JAM LAGI"
    elif level == "3h":
        icon, judul = "🚨", "PAKET AKAN EXPIRED — 3 JAM LAGI"
    else:
        icon, judul = "🔴", "PAKET AKAN EXPIRED — 1 JAM LAGI"

    msg = (
        f"{icon} <b>{judul}</b>\n\n"
        f"🏷️ Paket: {tier_emoji} <b>{tier_label}</b>\n"
        f"⏳ Sisa waktu: <b>{sisa_str}</b>\n\n"
        f"<blockquote>Segera perpanjang agar monitoring tidak terhenti!\n"
        f"Ketik /beli untuk pilihan paket.</blockquote>"
    )
    msg_owner = (
        f"{icon} <b>INFO PAKET USER</b>\n\n"
        f"👤 User ID: <code>{uid}</code>\n"
        f"🏷️ Paket: {tier_emoji} <b>{tier_label}</b>\n"
        f"⏳ Expired dalam: <b>{sisa_str}</b>"
    )
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data={"chat_id": uid, "text": msg, "parse_mode": "HTML"},
            timeout=10,
        )
    except Exception:
        pass
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data={"chat_id": OWNER_ID, "text": msg_owner, "parse_mode": "HTML"},
            timeout=10,
        )
    except Exception:
        pass


def run_expiry_notifier():
    """
    Background thread: cek expired paket premium setiap 30 menit.
    Kirim notif pada:
      - 24 jam sebelum expired
      -  3 jam sebelum expired
      -  1 jam sebelum expired
    Setelah expired, reset agar notif bisa terkirim lagi untuk perpanjangan berikutnya.
    """
    print(Fore.CYAN + "  EXPIRY NOTIFIER — background aktif (cek tiap 30 menit)")
    time.sleep(60)  # Tunggu bot ready

    while True:
        try:
            now = time.time()
            # Reload dari disk agar selalu data terbaru
            current_premium = load_premium()

            for uid_str, data in current_premium.items():
                expired_at = data.get("expired", 0)
                tier       = data.get("tier", "free")
                if tier == "free" or expired_at <= 0:
                    continue

                sisa = expired_at - now

                if sisa <= 0:
                    # Sudah expired — bersihkan state agar notif bisa jalan lagi setelah perpanjang
                    _notif_expiry_sent.pop(uid_str, None)
                    continue

                sent = _notif_expiry_sent.setdefault(uid_str, set())

                # 24 jam sebelum
                if sisa <= 86400 and "24h" not in sent:
                    sent.add("24h")
                    print(Fore.YELLOW + f"  EXPIRY WARN 24h: uid={uid_str} tier={tier}")
                    _send_expiry_notif(int(uid_str), tier, sisa, "24h")

                # 3 jam sebelum
                if sisa <= 10800 and "3h" not in sent:
                    sent.add("3h")
                    print(Fore.YELLOW + f"  EXPIRY WARN 3h: uid={uid_str} tier={tier}")
                    _send_expiry_notif(int(uid_str), tier, sisa, "3h")

                # 1 jam sebelum
                if sisa <= 3600 and "1h" not in sent:
                    sent.add("1h")
                    print(Fore.RED + f"  EXPIRY WARN 1h: uid={uid_str} tier={tier}")
                    _send_expiry_notif(int(uid_str), tier, sisa, "1h")

        except Exception as e:
            print(Fore.RED + f"ERROR EXPIRY NOTIFIER: {e}")

        time.sleep(1800)  # Cek ulang setiap 30 menit


def account_worker(acc):
    """
    Thread mandiri per akun — polling loop dengan adaptive sleep.
    - SMS ditemukan    → langsung poll lagi (sleep 1s)
    - Tidak ada SMS    → mundur bertahap hingga 5s (hemat CPU & koneksi)
    - Error / session  → backoff lebih panjang hingga 15s
    """
    email      = acc.get("email", "")
    sleep_time = 2.0   # interval awal
    while True:
        try:
            found = poll_one_account(acc)
            if found:
                sleep_time = 1.0                      # ada SMS — poll cepat
            else:
                sleep_time = min(sleep_time + 0.5, 5.0)  # naik pelan ke max 5s
        except Exception as e:
            print(Fore.RED + f"ERROR WORKER [{email}]: {e}")
            sleep_time = min(sleep_time * 2, 15.0)   # error → backoff agresif
        time.sleep(sleep_time)


# ================= BOT MANAGER (state sync + thread manager) =================
def run_bot():
    global _premium_acc_cache, _cache_dirty, _last_cache_save, _force_bot_sync
    _account_threads = {}   # email -> Thread
    _last_sync       = 0.0

    print(Fore.CYAN + Style.BRIGHT + "  BOT MANAGER STARTED — per-account threading aktif")

    while True:
        try:
            now = time.time()

            # ---- Sync state setiap 30 detik ATAU saat _force_bot_sync=True ----
            if now - _last_sync >= 30 or _force_bot_sync:

                # Rebuild email → user_id mapping
                new_email_to_uid = {}
                with accounts_lock:
                    for acc in accounts:
                        new_email_to_uid[acc["email"]] = OWNER_ID

                users_data = load_users()
                owner_emails = set(new_email_to_uid.keys())
                for uid_str, udata in users_data.items():
                    try:
                        uid_int = int(uid_str)
                    except Exception:
                        continue
                    for em in udata.get("emails", []):
                        if em not in owner_emails:
                            new_email_to_uid[em] = uid_int

                # Sync premium account sessions (cookie-based + password-based user accounts)
                prem_cookies = load_premium_cookies()
                active_prem_emails = set()
                for uid_str, udata in users_data.items():
                    # --- Cookie-based (legacy / manual cookie) ---
                    for em in udata.get("emails", []):
                        if em in owner_emails or em not in prem_cookies:
                            continue
                        active_prem_emails.add(em)
                        if em not in _premium_acc_cache:
                            prem_acc = {
                                "email": em, "password": None,
                                "cookies": prem_cookies[em],
                                "session": make_httpx_client(),
                                "last_login": 0, "csrf_token": "",
                            }
                            prem_acc["session"].cookies.update(prem_cookies[em])
                            _premium_acc_cache[em] = prem_acc
                        else:
                            cached = _premium_acc_cache[em]
                            if cached.get("cookies") != prem_cookies[em]:
                                cached["cookies"] = prem_cookies[em]
                                cached["session"].cookies.clear()
                                cached["session"].cookies.update(prem_cookies[em])
                                cached["last_login"] = 0

                    # --- Password-based (via /addemail email password) ---
                    for ua in udata.get("user_accounts", []):
                        em  = ua.get("email", "")
                        pwd = ua.get("password", "")
                        if not em or not pwd or em in owner_emails:
                            continue
                        active_prem_emails.add(em)
                        if em not in _premium_acc_cache:
                            prem_acc = {
                                "email": em, "password": pwd,
                                "cookies": {}, "session": make_httpx_client(),
                                "last_login": 0, "csrf_token": "",
                            }
                            _premium_acc_cache[em] = prem_acc
                        else:
                            # Update password jika berubah
                            _premium_acc_cache[em]["password"] = pwd

                for em in list(_premium_acc_cache.keys()):
                    if em not in active_prem_emails:
                        del _premium_acc_cache[em]

                # Update shared state (atomic dict replace)
                with accounts_lock:
                    all_accs = list(accounts) + list(_premium_acc_cache.values())
                _bot_state["email_to_uid"]   = new_email_to_uid
                _bot_state["total_accounts"] = len(all_accs)

                # Spawn thread baru untuk akun yang belum punya thread / thread mati
                active_emails = set()
                for acc in all_accs:
                    em = acc.get("email", "")
                    if not em:
                        continue
                    active_emails.add(em)
                    t = _account_threads.get(em)
                    if t is None or not t.is_alive():
                        nt = threading.Thread(
                            target=account_worker, args=(acc,),
                            daemon=True, name=f"poll-{em[:25]}"
                        )
                        nt.start()
                        _account_threads[em] = nt
                        print(Fore.YELLOW + f"  THREAD START: {em}")

                # Hapus thread untuk akun yang sudah dihapus
                for em in [e for e in list(_account_threads) if e not in active_emails]:
                    del _account_threads[em]

                _last_sync       = now
                _force_bot_sync  = False  # Reset flag setelah sync selesai

            else:
                # Health-check ringan setiap siklus (2 detik) — respawn thread mati
                for em, t in list(_account_threads.items()):
                    if not t.is_alive():
                        print(Fore.YELLOW + f"  THREAD MATI, RESPAWN: {em}")
                        with accounts_lock:
                            all_now = list(accounts) + list(_premium_acc_cache.values())
                        for acc in all_now:
                            if acc.get("email") == em:
                                nt = threading.Thread(
                                    target=account_worker, args=(acc,),
                                    daemon=True, name=f"poll-{em[:25]}"
                                )
                                nt.start()
                                _account_threads[em] = nt
                                break

            # Flush cache kalau dirty tapi belum sempat tersimpan
            if _cache_dirty and time.time() - _last_cache_save >= 5:
                with _sent_cache_lock:
                    save_sent_cache()
                _last_cache_save = time.time()
                _cache_dirty = False

            time.sleep(2)

        except Exception as e:
            print(Fore.RED + f"ERROR BOT MANAGER: {e}")
            time.sleep(2)

            
# ================= KEEP-ALIVE SERVER =================
import json as _json
from http.server import HTTPServer, BaseHTTPRequestHandler

_bot_start_time = time.time()

class KeepAliveHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        path = self.path.split("?")[0].rstrip("/")

        if path in ("", "/", "/health"):
            # Endpoint ringan untuk UptimeRobot / Railway health check
            body = b"OK"
            self.send_response(200)
            self.send_header("Content-Type", "text/plain")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        elif path == "/status":
            # Endpoint JSON lengkap — untuk monitoring manual
            now = time.time()
            uptime_sec = int(now - _bot_start_time)
            uptime_str = f"{uptime_sec // 3600}h {(uptime_sec % 3600) // 60}m {uptime_sec % 60}s"

            with accounts_lock:
                owner_list = [
                    {
                        "email": a.get("email", ""),
                        "active": now - a.get("last_login", 0) < LOGIN_COOLDOWN,
                        "last_keepalive": int(now - _last_cookie_refresh.get(a.get("email",""), 0))
                    }
                    for a in accounts
                ]
            prem_list = [
                {
                    "email": e,
                    "active": now - a.get("last_login", 0) < LOGIN_COOLDOWN,
                    "last_keepalive": int(now - _last_cookie_refresh.get(e, 0))
                }
                for e, a in _premium_acc_cache.items()
            ]

            data = {
                "status": "running",
                "uptime": uptime_str,
                "uptime_seconds": uptime_sec,
                "owner_accounts": owner_list,
                "premium_accounts": prem_list,
                "total_accounts": len(owner_list) + len(prem_list),
                "keepalive_interval_sec": COOKIE_KEEPALIVE_INTERVAL,
            }
            body = _json.dumps(data, indent=2).encode()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        else:
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b"Not found")

    def log_message(self, format, *args):
        pass  # Nonaktifkan log HTTP agar console bersih

def run_keepalive():
    port = int(os.environ.get("PORT", 5000))
    HTTPServer.allow_reuse_address = True
    server = HTTPServer(("0.0.0.0", port), KeepAliveHandler)
    print(Fore.CYAN + f"  KEEP-ALIVE SERVER — port {port} | /health /status")
    server.serve_forever()

# ================= GRACEFUL SHUTDOWN (Railway SIGTERM) =================
def _graceful_shutdown(signum, frame):
    print(Fore.YELLOW + "\n  SIGNAL DITERIMA — menyimpan state sebelum shutdown...")
    try:
        with _sent_cache_lock:
            save_sent_cache()
    except Exception:
        pass
    print(Fore.YELLOW + "  State tersimpan. Bye!")
    sys.exit(0)

signal.signal(signal.SIGTERM, _graceful_shutdown)
signal.signal(signal.SIGINT,  _graceful_shutdown)

# ================= START BOT =================
# Ambil username bot via getMe untuk link tombol laporan
def _init_bot_username():
    global BOT_USERNAME
    try:
        r = requests.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getMe", timeout=10)
        d = r.json()
        if d.get("ok"):
            BOT_USERNAME = d["result"].get("username", "")
            print(Fore.CYAN + f"  BOT USERNAME: @{BOT_USERNAME}")
    except Exception as e:
        print(Fore.YELLOW + f"  getMe error: {e}")
_init_bot_username()

threading.Thread(target=run_keepalive,        daemon=True).start()
threading.Thread(target=listen_command,       daemon=True).start()
threading.Thread(target=auto_cookie_refresher,daemon=True).start()
threading.Thread(target=run_auto_backup,      daemon=True).start()
threading.Thread(target=run_expiry_notifier,  daemon=True).start()
run_bot()
